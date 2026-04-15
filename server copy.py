#...existing code...
# ══════════════════════════════════════════════════════════════════════════════
# API: Create Jira Task from Recommendation
# ══════════════════════════════════════════════════════════════════════════════

import os
import re
import json
import base64
import uuid
import logging
import time
import hashlib
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Any, Optional
from functools import wraps

from flask import (
    Flask, render_template, request, jsonify,
    send_file, redirect, session, url_for, abort
)
from flask_cors import CORS
from dotenv import load_dotenv
from authlib.integrations.flask_client import OAuth

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
import requests
from requests.auth import HTTPBasicAuth

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage

from mcp_client import call_jira_tool, list_jira_tools, MCPClientError
from flask_session import Session

# ── Env + logging ─────────────────────────────────────────────────────────────
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

JIRA_OAUTH_SCOPES = "read:jira-work write:jira-work read:jira-user read:board-scope:jira-software read:sprint:jira-software offline_access"
#JIRA_OAUTH_SCOPES = "read:jira-work write:jira-work read:jira-user read:board-scope:jira-software read:sprint:jira-software offline_access"JIRA_OAUTH_SCOPES = "read:jira-work write:jira-work read:jira-user offline_access"

# ── Flask ─────────────────────────────────────────────────────────────────────
app = Flask(__name__)
# IMPORTANT: Must be a fixed key — os.urandom() changes on restart and breaks OAuth state
_default_secret = "jira-dashboard-secret-key-change-this-in-production-2026"
app.secret_key = os.getenv("FLASK_SECRET_KEY", _default_secret)
app.config["PERMANENT_SESSION_LIFETIME"] = 28800  # 8 hours
app.config["SESSION_TYPE"]            = "filesystem"
app.config["SESSION_FILE_DIR"]        = "/tmp/jira_dashboard_sessions"
app.config["SESSION_FILE_THRESHOLD"]  = 100
CORS(app)

# Session config — critical for OAuth state to survive the redirect round-trip
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"    # allow cross-site redirect callbacks
app.config["SESSION_COOKIE_SECURE"]   = False     # False for localhost (True in production)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_NAME"]     = "jira_dashboard_session" 



# ── OAuth ─────────────────────────────────────────────────────────────────────
# Init server-side session (stores OAuth state server-side, not in cookie)
import os as _os
_os.makedirs("/tmp/jira_dashboard_sessions", exist_ok=True)
Session(app)

oauth = OAuth(app)
atlassian = oauth.register(
    name="atlassian",
    client_id=os.getenv("JIRA_CLIENT_ID"),
    client_secret=os.getenv("JIRA_CLIENT_SECRET"),
    authorize_url="https://auth.atlassian.com/authorize",
    access_token_url="https://auth.atlassian.com/oauth/token",
    client_kwargs={
        "scope": JIRA_OAUTH_SCOPES,
        "audience": "api.atlassian.com",
        "prompt": "consent",
    },
)

# ── Per-user dashboard store: {user_id: {dash_id: dash_data}} ────────────────
_user_dashboards: Dict[str, Dict] = {}

# ── Per-user Jira cache: {cache_key: (data, timestamp)} ─────────────────────
_user_cache: Dict[str, Any] = {}
CACHE_TTL = 60


# ══════════════════════════════════════════════════════════════════════════════
# Session helpers — everything scoped to current user
# ══════════════════════════════════════════════════════════════════════════════

def current_user() -> Optional[Dict]:
    """Return current logged-in user info from session."""
    return session.get("user")


def user_id() -> Optional[str]:
    """Return a stable user ID (hashed email)."""
    user = current_user()
    if not user:
        return None
    return hashlib.md5(user.get("email", "").encode()).hexdigest()


def is_authenticated() -> bool:
    """Check if user is logged in and token exists."""
    return bool(session.get("oauth_token") and session.get("jira_cloud_id") and current_user())


def get_user_project_key() -> str:
    """Get the Jira project key selected by the current user."""
    return session.get("project_key", "")


def get_user_dashboards() -> Dict:
    """Get dashboards for the current user only."""
    uid = user_id()
    if not uid:
        return {}
    return _user_dashboards.get(uid, {})


def save_user_dashboard(dash_id: str, data: Dict):
    """Save a dashboard for the current user."""
    uid = user_id()
    if not uid:
        return
    if uid not in _user_dashboards:
        _user_dashboards[uid] = {}
    _user_dashboards[uid][dash_id] = data


# ── login_required decorator ──────────────────────────────────────────────────
def login_required(f):
    """Redirect to login if user not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not is_authenticated():
            if request.path.startswith("/api/") or request.is_json:
                return jsonify({"error": "Not authenticated", "redirect": "/login"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
# Jira Auth — uses current user's OAuth token
# ══════════════════════════════════════════════════════════════════════════════

def get_jira_auth() -> Optional[Dict]:
    """
    Return Jira connection config for the CURRENT USER.
    Uses their OAuth token from session.
    Falls back to .env API key only if no OAuth token present.
    """
    token    = session.get("oauth_token")
    cloud_id = session.get("jira_cloud_id")

    if token and cloud_id:
        return {
            "base_url":  f"https://api.atlassian.com/ex/jira/{cloud_id}",
            "auth":      None,
            "headers":   {
                "Authorization": f"Bearer {token.get('access_token')}",
                "Accept":        "application/json",
                "Content-Type":  "application/json",
            },
            "auth_type": "oauth",
            "site":      session.get("jira_site_name", ""),
        }

    # API key fallback
    server    = os.getenv("JIRA_SERVER", "").rstrip("/")
    email     = os.getenv("JIRA_USER_EMAIL")
    api_token = os.getenv("JIRA_API_TOKEN")
    if all([server, email, api_token]):
        return {
            "base_url":  server,
            "auth":      HTTPBasicAuth(email, api_token),
            "headers":   {"Accept": "application/json", "Content-Type": "application/json"},
            "auth_type": "apikey",
            "site":      server.replace("https://", ""),
        }

    return None


def get_mcp_auth_context() -> Optional[Dict]:
    """Return session-scoped auth context passed to MCP tool calls."""
    token = session.get("oauth_token")
    cloud_id = session.get("jira_cloud_id")

    if token and cloud_id:
        return {
            "auth_type": "oauth",
            "cloud_id": cloud_id,
            "access_token": token.get("access_token", ""),
            "site": session.get("jira_site_name", ""),
        }

    server = os.getenv("JIRA_SERVER", "").rstrip("/")
    email = os.getenv("JIRA_USER_EMAIL")
    api_token = os.getenv("JIRA_API_TOKEN")
    if all([server, email, api_token]):
        return {
            "auth_type": "apikey",
            "server": server,
            "email": email,
            "api_token": api_token,
        }

    return None


def jira_get(path: str, params: dict = None) -> dict:
    cfg  = get_jira_auth()
    if not cfg:
        raise ValueError("Not authenticated with Jira")
    logger.info(f"[{cfg['auth_type'].upper()}] GET {path}")
    url  = f"{cfg['base_url']}/rest/api/3/{path.lstrip('/')}"
    resp = requests.get(url, auth=cfg.get("auth"), headers=cfg["headers"], params=params)
    resp.raise_for_status()
    return resp.json()


def jira_get_cached(path: str, params: dict = None) -> dict:
    """Cached version of jira_get — 60s TTL, scoped per user."""
    uid = user_id() or "anon"
    key = f"{uid}:{path}:{json.dumps(params or {}, sort_keys=True)}"
    now = time.time()
    if key in _user_cache:
        data, ts = _user_cache[key]
        if now - ts < CACHE_TTL:
            logger.info(f"[CACHE HIT] {path}")
            return data
    data = jira_get(path, params)
    _user_cache[key] = (data, now)
    return data


def jira_post(path: str, payload: dict) -> dict:
    cfg  = get_jira_auth()
    if not cfg:
        raise ValueError("Not authenticated with Jira")
    logger.info(f"[{cfg['auth_type'].upper()}] POST {path}")
    url  = f"{cfg['base_url']}/rest/api/3/{path.lstrip('/')}"
    resp = requests.post(url, auth=cfg.get("auth"), headers=cfg["headers"], json=payload)
    resp.raise_for_status()
    return resp.json() if resp.content else {}


def jira_put(path: str, payload: dict) -> dict:
    cfg  = get_jira_auth()
    if not cfg:
        raise ValueError("Not authenticated with Jira")
    url  = f"{cfg['base_url']}/rest/api/3/{path.lstrip('/')}"
    resp = requests.put(url, auth=cfg.get("auth"), headers=cfg["headers"], json=payload)
    resp.raise_for_status()
    return resp.json() if resp.content else {}


def clear_user_cache():
    """Clear cache entries for the current user."""
    uid = user_id() or "anon"
    keys_to_delete = [k for k in _user_cache if k.startswith(f"{uid}:")]
    for k in keys_to_delete:
        del _user_cache[k]
    logger.info(f"Cache cleared for user {uid}")


def parse_jira_datetime(value: str) -> Optional[datetime]:
    """Parse Jira datetime strings into timezone-aware datetime objects."""
    if not value:
        return None
    try:
        normalized = value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def is_done_status(status_name: str) -> bool:
    status = (status_name or "").strip().lower()
    return status in {"done", "closed", "resolved", "complete", "completed"}


def is_blocked_issue(fields: Dict[str, Any]) -> bool:
    status_name = (fields.get("status") or {}).get("name", "")
    labels = [str(label).lower() for label in fields.get("labels", [])]
    summary = (fields.get("summary") or "").lower()
    blocked_terms = ["blocked", "blocker", "impediment"]
    return any(term in status_name.lower() for term in blocked_terms) or any(term in labels for term in blocked_terms) or any(term in summary for term in blocked_terms)


def clamp_score(value: float, lower: int = 0, upper: int = 100) -> int:
    return max(lower, min(upper, int(round(value))))


def llm_synthesize_project_intelligence(base_intelligence: Dict[str, Any]) -> Dict[str, Any]:
    """Use LLM to generate decision-focused narrative and recommendations."""
    model = llm_fast or llm
    if not model:
        return {}

    system_prompt = """You are a senior delivery intelligence analyst for software projects.
Given project delivery metrics, return ONLY strict JSON with this schema:
{
  "executive_summary": "string",
  "health_score": 0,
  "health_label": "Good|Watch|Risk",
  "risk_level": "Low|Medium|High",
  "forecast_summary": "string",
  "recommendations": [
    {"title": "string", "reason": "string", "impact": "string"}
  ]
}

Rules:
- recommendations should contain 5 to 8 concise, action-oriented items when possible.
- health_score must be an integer 0..100.
- Be concrete and managerial; avoid vague language.
- Return JSON only, no markdown.
"""

    compact_input = {
        "project_key": base_intelligence.get("project_key"),
        "indicators": base_intelligence.get("indicators", {}),
        "forecast": {
            "recent_completed": base_intelligence.get("forecast", {}).get("recent_completed"),
            "recent_created": base_intelligence.get("forecast", {}).get("recent_created"),
            "completion_rate": base_intelligence.get("forecast", {}).get("completion_rate"),
            "estimated_days_to_clear": base_intelligence.get("forecast", {}).get("estimated_days_to_clear"),
            "summary": base_intelligence.get("forecast", {}).get("summary"),
        },
        "highlights": base_intelligence.get("highlights", []),
        "current_health": {
            "score": base_intelligence.get("health_score"),
            "label": base_intelligence.get("health_label"),
            "risk_level": base_intelligence.get("risk_level"),
        },
    }

    try:
        response = model.invoke([
            SystemMessage(content=system_prompt),
            HumanMessage(content=json.dumps(compact_input))
        ])
        raw = (response.content or "").strip()
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if not match:
            return {}
        parsed = json.loads(match.group())

        out: Dict[str, Any] = {}
        score = parsed.get("health_score")
        if isinstance(score, (int, float)):
            out["health_score"] = clamp_score(score)

        label = str(parsed.get("health_label", "")).strip().title()
        if label in {"Good", "Watch", "Risk"}:
            out["health_label"] = label

        risk_level = str(parsed.get("risk_level", "")).strip().title()
        if risk_level in {"Low", "Medium", "High"}:
            out["risk_level"] = risk_level

        executive_summary = str(parsed.get("executive_summary", "")).strip()
        if executive_summary:
            out["executive_summary"] = executive_summary[:500]

        forecast_summary = str(parsed.get("forecast_summary", "")).strip()
        if forecast_summary:
            out.setdefault("forecast", {})
            out["forecast"]["summary"] = forecast_summary[:400]

        recommendations = parsed.get("recommendations", [])
        clean_recs = []
        if isinstance(recommendations, list):
            for item in recommendations[:8]:
                if not isinstance(item, dict):
                    continue
                title = str(item.get("title", "")).strip()
                reason = str(item.get("reason", "")).strip()
                impact = str(item.get("impact", "")).strip()
                if title and reason and impact:
                    clean_recs.append({"title": title[:120], "reason": reason[:280], "impact": impact[:220]})
        if clean_recs:
            out["recommendations"] = clean_recs

        return out
    except Exception as e:
        logger.warning(f"LLM intelligence synthesis failed: {e}")
        return {}


def llm_generate_assignment_suggestions(
    project_key: str,
    unassigned_issues: List[Dict[str, Any]],
    assignee_loads: List[Dict[str, Any]],
) -> List[Dict[str, str]]:
    """Generate assignee recommendations for unassigned issues using LLM."""
    model = llm_fast or llm
    if not model or not unassigned_issues or not assignee_loads:
        return []

    allowed_assignees = [str(a.get("name", "")).strip() for a in assignee_loads if str(a.get("name", "")).strip()]
    if not allowed_assignees:
        return []

    compact_input = {
        "project_key": project_key,
        "unassigned_issues": [
            {
                "key": i.get("key", ""),
                "summary": i.get("summary", ""),
                "priority": i.get("priority", ""),
                "status": i.get("status", ""),
            }
            for i in unassigned_issues[:8]
        ],
        "assignee_loads": assignee_loads[:10],
    }

    system_prompt = """You assign unowned Jira issues to team members.
Return ONLY strict JSON in this schema:
{
  "suggestions": [
    {
      "key": "ISSUE-1",
      "primary": "Assignee Name",
      "backup": "Assignee Name",
      "reason": "short reason"
    }
  ]
}

Rules:
- Use only assignee names from provided assignee_loads.
- Prefer lower-load assignees for primary assignments.
- backup can equal primary when only one assignee exists.
- Keep reason concise (max 140 chars).
- JSON only, no markdown.
"""

    try:
        response = model.invoke([
            SystemMessage(content=system_prompt),
            HumanMessage(content=json.dumps(compact_input))
        ])
        raw = (response.content or "").strip()
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if not match:
            return []
        parsed = json.loads(match.group())
        suggestions = parsed.get("suggestions", [])
        if not isinstance(suggestions, list):
            return []

        valid_keys = {str(i.get("key", "")) for i in unassigned_issues}
        allowed_set = set(allowed_assignees)
        clean: List[Dict[str, str]] = []
        for item in suggestions[:8]:
            if not isinstance(item, dict):
                continue
            key = str(item.get("key", "")).strip()
            primary = str(item.get("primary", "")).strip()
            backup = str(item.get("backup", "")).strip() or primary
            reason = str(item.get("reason", "")).strip()
            if key not in valid_keys:
                continue
            if primary not in allowed_set:
                continue
            if backup not in allowed_set:
                backup = primary
            clean.append({
                "key": key,
                "primary": primary,
                "backup": backup,
                "reason": reason[:140],
            })

        return clean
    except Exception as e:
        logger.warning(f"LLM assignment suggestion failed: {e}")
        return []


def build_project_intelligence(project_key: str) -> Dict[str, Any]:
    """Compute project health, forecast, and recommended actions from Jira issues."""
    if not project_key:
        raise ValueError("Project key is required")
    if not get_jira_auth():
        raise ValueError("Jira not connected")

    open_issues_data = jira_get_cached("search/jql", params={
        "jql": f"project={project_key} ORDER BY updated DESC",
        "maxResults": 200,
        "fields": "summary,status,assignee,priority,created,updated,duedate,labels,issuetype",
    })
    recent_done_data = jira_get_cached("search/jql", params={
        "jql": f"project={project_key} AND statusCategory = Done AND resolved >= -14d ORDER BY updated DESC",
        "maxResults": 1,
        "fields": "resolutiondate",
    })
    recent_created_data = jira_get_cached("search/jql", params={
        "jql": f"project={project_key} AND created >= -14d ORDER BY created DESC",
        "maxResults": 1,
        "fields": "created",
    })

    raw_issues = open_issues_data.get("issues", [])
    issues = []
    now = datetime.now(timezone.utc)
    assignee_counts: Dict[str, int] = {}
    overdue_issues = 0
    blocked_issues = 0
    high_priority_open = 0
    unassigned_open = 0
    stale_in_progress = 0

    for issue in raw_issues:
        fields = issue.get("fields", {})
        status_name = (fields.get("status") or {}).get("name", "")
        if is_done_status(status_name):
            continue

        assignee_name = (fields.get("assignee") or {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned"
        priority_name = (fields.get("priority") or {}).get("name", "") if fields.get("priority") else ""
        updated_at = parse_jira_datetime(fields.get("updated", ""))
        due_date = parse_jira_datetime(fields.get("duedate", ""))
        if due_date and due_date < now:
            overdue_issues += 1
        if assignee_name == "Unassigned":
            unassigned_open += 1
        if priority_name.lower() in {"high", "highest", "critical", "blocker"}:
            high_priority_open += 1
        if is_blocked_issue(fields):
            blocked_issues += 1
        if updated_at and (now - updated_at) > timedelta(days=7) and status_name.lower() in {"in progress", "in review", "review", "testing"}:
            stale_in_progress += 1

        assignee_counts[assignee_name] = assignee_counts.get(assignee_name, 0) + 1
        issues.append({
            "key": issue.get("key", ""),
            "summary": fields.get("summary", ""),
            "status": status_name,
            "assignee": assignee_name,
            "priority": priority_name,
        })

    open_count = len(issues)
    recent_done = recent_done_data.get("total", 0)
    recent_created = recent_created_data.get("total", 0)
    highest_load = max(assignee_counts.values()) if assignee_counts else 0
    most_loaded_assignee = max(assignee_counts, key=assignee_counts.get) if assignee_counts else None

    score = 100.0
    if open_count:
        score -= (overdue_issues / open_count) * 30
        score -= (blocked_issues / open_count) * 22
        score -= (stale_in_progress / open_count) * 18
        score -= (unassigned_open / open_count) * 10
        if highest_load >= 5 and highest_load / open_count > 0.35:
            score -= 10
    if recent_created > recent_done:
        score -= min(15, (recent_created - recent_done) * 2)

    health_score = clamp_score(score)
    if health_score >= 80:
        health_label = "Good"
        risk_level = "Low"
    elif health_score >= 60:
        health_label = "Watch"
        risk_level = "Medium"
    else:
        health_label = "Risk"
        risk_level = "High"

    completion_rate = round((recent_done / max(recent_created, 1)) * 100, 1) if recent_created else 100.0
    estimated_days_to_clear = None
    if recent_done > 0 and open_count > 0:
        estimated_days_to_clear = round((open_count / max(recent_done / 14, 0.1)), 1)

    forecast_parts = []
    if recent_created > recent_done:
        forecast_parts.append("Incoming work is outpacing completions")
    if blocked_issues:
        forecast_parts.append(f"{blocked_issues} blocker(s) may delay delivery")
    if estimated_days_to_clear:
        forecast_parts.append(f"Backlog clearance estimate is {estimated_days_to_clear} days")
    if not forecast_parts:
        forecast_parts.append("Current delivery trend looks stable")

    recommendations = []
    if blocked_issues:
        recommendations.append({
            "title": "Resolve blockers first",
            "reason": f"{blocked_issues} open blocker(s) are increasing delivery risk.",
            "impact": "Removes immediate schedule risk.",
        })
    if highest_load >= 5 and most_loaded_assignee:
        recommendations.append({
            "title": f"Rebalance work from {most_loaded_assignee}",
            "reason": f"{most_loaded_assignee} owns {highest_load} active issue(s).",
            "impact": "Reduces bottlenecks and improves flow.",
        })
    if stale_in_progress:
        recommendations.append({
            "title": "Review stale in-progress work",
            "reason": f"{stale_in_progress} issue(s) have been inactive for over 7 days.",
            "impact": "Improves delivery predictability.",
        })
    if unassigned_open:
        recommendations.append({
            "title": "Assign unowned issues",
            "reason": f"{unassigned_open} issue(s) have no owner.",
            "impact": "Clarifies accountability and next steps.",
        })
    if not recommendations:
        recommendations.append({
            "title": "Maintain current delivery pace",
            "reason": "No major operational risks detected in the latest issue set.",
            "impact": "Keeps the project on track.",
        })

    highlights = [
        f"{open_count} open issue(s)",
        f"{recent_done} completed in the last 14 days",
        f"{recent_created} created in the last 14 days",
        f"{completion_rate}% completion-to-intake ratio",
    ]

    active_assignees = [
        {"name": name, "count": count}
        for name, count in assignee_counts.items()
        if str(name).strip().lower() != "unassigned"
    ]
    active_assignees.sort(key=lambda a: a["count"])

    unassigned_issues = [i for i in issues if str(i.get("assignee", "")).strip().lower() == "unassigned"]

    rule_assignment_suggestions: List[Dict[str, str]] = []
    if unassigned_issues and active_assignees:
        for idx, issue in enumerate(unassigned_issues[:8]):
            primary = active_assignees[idx % len(active_assignees)]
            backup = active_assignees[(idx + 1) % len(active_assignees)] if len(active_assignees) > 1 else primary
            rule_assignment_suggestions.append({
                "key": issue.get("key", ""),
                "primary": primary.get("name", ""),
                "backup": backup.get("name", ""),
                "reason": "Balanced by current assignee load.",
            })

    intelligence = {
        "project_key": project_key,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "health_score": health_score,
        "health_label": health_label,
        "risk_level": risk_level,
        "forecast": {
            "recent_completed": recent_done,
            "recent_created": recent_created,
            "completion_rate": completion_rate,
            "estimated_days_to_clear": estimated_days_to_clear,
            "summary": ". ".join(forecast_parts),
        },
        "indicators": {
            "open_issues": open_count,
            "overdue_issues": overdue_issues,
            "blocked_issues": blocked_issues,
            "high_priority_open": high_priority_open,
            "unassigned_open": unassigned_open,
            "stale_in_progress": stale_in_progress,
        },
        "highlights": highlights,
        "recommendations": recommendations[:8],
        "assignment_suggestions": rule_assignment_suggestions,
        "assignment_suggestions_source": "rules",
        "executive_summary": " ".join(highlights[:2]),
        "analysis_source": "rules",
    }

    llm_intel = llm_synthesize_project_intelligence(intelligence)
    if llm_intel:
        if "health_score" in llm_intel:
            intelligence["health_score"] = llm_intel["health_score"]
        if "health_label" in llm_intel:
            intelligence["health_label"] = llm_intel["health_label"]
        if "risk_level" in llm_intel:
            intelligence["risk_level"] = llm_intel["risk_level"]
        if "executive_summary" in llm_intel:
            intelligence["executive_summary"] = llm_intel["executive_summary"]
        if "forecast" in llm_intel and isinstance(llm_intel["forecast"], dict):
            intelligence["forecast"].update(llm_intel["forecast"])
        if "recommendations" in llm_intel:
            intelligence["recommendations"] = llm_intel["recommendations"]
        intelligence["analysis_source"] = "llm"

    llm_assignment = llm_generate_assignment_suggestions(project_key, unassigned_issues, active_assignees)
    if llm_assignment:
        intelligence["assignment_suggestions"] = llm_assignment
        intelligence["assignment_suggestions_source"] = "llm"

    return intelligence


# ══════════════════════════════════════════════════════════════════════════════
# Jira site helpers
# ══════════════════════════════════════════════════════════════════════════════

def fetch_accessible_sites(access_token: str) -> List[Dict]:
    """Fetch all Jira sites the user has access to."""
    try:
        resp = requests.get(
            "https://api.atlassian.com/oauth/token/accessible-resources",
            headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.error(f"fetch_accessible_sites failed: {e}")
        return []


def fetch_user_projects(cloud_id: str, access_token: str) -> List[Dict]:
    """Fetch all projects from the user's Jira site."""
    try:
        url  = f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/api/3/project"
        resp = requests.get(
            url,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Accept":        "application/json",
            }
        )
        resp.raise_for_status()
        projects = resp.json()
        return [{"key": p["key"], "name": p["name"]} for p in projects]
    except Exception as e:
        logger.error(f"fetch_user_projects failed: {e}")
        return []


def ensure_selected_project() -> bool:
    """Ensure a valid project is selected in session for the current Jira site."""
    token = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    access_token = token.get("access_token", "")
    if not cloud_id or not access_token:
        return False

    projects = fetch_user_projects(cloud_id, access_token)
    if not projects:
        return False

    current_key = (session.get("project_key") or "").upper().strip()
    current = next((p for p in projects if (p.get("key") or "").upper() == current_key), None)

    if current:
        session["project_key"] = current.get("key", "")
        session["project_name"] = current.get("name", "")
        return True

    first = projects[0]
    session["project_key"] = first.get("key", "")
    session["project_name"] = first.get("name", "")
    return True


# ══════════════════════════════════════════════════════════════════════════════
# LLM setup
# ══════════════════════════════════════════════════════════════════════════════

def get_llm(fast: bool = False) -> Optional[ChatOpenAI]:
    api_key  = os.getenv("QWEN_API_KEY", "")
    base_url = os.getenv("QWEN_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
    model    = os.getenv("QWEN_MODEL_FAST" if fast else "QWEN_MODEL", "qwen-plus")

    if not api_key:
        return None
    try:
        return ChatOpenAI(
            model=model, temperature=0.1,
            api_key=api_key, base_url=base_url,
            request_timeout=30, max_retries=1,
        )
    except Exception as e:
        logger.error(f"LLM init failed: {e}")
        return None


llm      = get_llm(fast=False)
llm_fast = get_llm(fast=True)


# ══════════════════════════════════════════════════════════════════════════════
# Fast dispatch — skip LLM for obvious queries
# ══════════════════════════════════════════════════════════════════════════════

def fast_dispatch(prompt_lower: str, project_key: str) -> Optional[Dict]:
    auth_context = get_mcp_auth_context()
    if not auth_context:
        return None

    def _mcp_call(name: str, args: Dict[str, Any]) -> Dict:
        try:
            merged = {"auth_context": auth_context, **args}
            return call_jira_tool(name, merged)
        except MCPClientError as e:
            logger.warning(f"FAST MCP call failed ({name}): {e}")
            if name == "jira_search_issues":
                try:
                    jql = args.get("jql", f"project={project_key} ORDER BY created DESC")
                    max_results = int(args.get("max_results", 50) or 50)
                    data = jira_get_cached("search/jql", params={
                        "jql": jql,
                        "maxResults": max_results,
                        "fields": "summary,status,assignee,priority,created,updated,issuetype",
                    })
                    raw_issues = data.get("issues", [])
                    issues = []
                    for item in raw_issues:
                        fields = item.get("fields", {})
                        issues.append({
                            "key": item.get("key", ""),
                            "summary": fields.get("summary", ""),
                            "status": fields.get("status", {}).get("name", ""),
                            "assignee": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
                            "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
                            "type": fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else "",
                            "created": fields.get("created", "")[:10] if fields.get("created") else "",
                            "updated": fields.get("updated", "")[:10] if fields.get("updated") else "",
                        })
                    return {"issues": issues, "total": data.get("total", len(issues))}
                except Exception as fallback_err:
                    logger.warning(f"FAST REST fallback failed (jira_search_issues): {fallback_err}")
            return {"error": str(e)}

    if any(p in prompt_lower for p in ["list sprint", "show sprint", "get sprint", "all sprint", "sprints"]):
        logger.info("[FAST] jira_get_sprints")
        result = _mcp_call("jira_get_sprints", {"project_key": project_key})
        if result.get("error"):
            return {"error": result["error"], "response": result["error"]}
        sprints = result.get("sprints", [])
        return {"sprints": sprints, "result_type": "sprints",
                "response": f"Found {len(sprints)} sprint(s)", "jira_results": sprints}

    if any(p in prompt_lower for p in ["my issue", "my task", "assigned to me", "what am i working"]):
        logger.info("[FAST] jira_get_my_issues")
        result = _mcp_call("jira_get_my_issues", {"project_key": project_key})
        issues = result.get("issues", [])
        return {"issues": issues, "result_type": "issues",
                "response": f"Found {len(issues)} issue(s) assigned to you", "jira_results": issues}

    issue_match = re.search(r"[A-Z]+-[0-9]+", prompt_lower.upper())
    if issue_match and any(p in prompt_lower for p in ["tell me", "details", "about", "show", "status of"]):
        logger.info(f"[FAST] jira_get_issue({issue_match.group()})")
        result = _mcp_call("jira_get_issue", {"issue_key": issue_match.group()})
        return {"result_type": "issue_detail", "jira_results": [result],
                "response": f"Details for {issue_match.group()}"}

    if any(p in prompt_lower for p in ["list issue", "show issue", "all issue", "show all", "list all"]):
        # Only treat "in <project>" as a project hint when it appears at the end,
        # so phrases like "in progress" do not get parsed as project keys.
        project_hint_match = re.search(r"\bin\s+([a-z][a-z0-9]+)\s*$", prompt_lower)
        effective_project_key = project_key
        if project_hint_match:
            hinted = project_hint_match.group(1).upper().strip()
            if 1 <= len(hinted) <= 12:
                effective_project_key = hinted

        sprint_match = re.search(r"sprint\s*(\d+)", prompt_lower)
        jql = f"project={effective_project_key}"
        if sprint_match:
            jql += f' AND sprint = "{effective_project_key} Sprint {sprint_match.group(1)}"'
        if "in progress" in prompt_lower:   jql += ' AND status = "In Progress"'
        elif "to do"      in prompt_lower:  jql += ' AND status = "To Do"'
        elif "done"       in prompt_lower:  jql += ' AND status = "Done"'
        jql += " ORDER BY created DESC"
        logger.info(f"[FAST] direct jira_search_issues({jql})")
        try:
            data = jira_get_cached("search/jql", params={
                "jql": jql,
                "maxResults": 50,
                "fields": "summary,status,assignee,priority,created,updated,issuetype",
            })
            raw_issues = data.get("issues", [])
            issues = []
            for item in raw_issues:
                fields = item.get("fields", {})
                issues.append({
                    "key": item.get("key", ""),
                    "summary": fields.get("summary", ""),
                    "status": fields.get("status", {}).get("name", ""),
                    "assignee": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
                    "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
                    "type": fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else "",
                    "created": fields.get("created", "")[:10] if fields.get("created") else "",
                    "updated": fields.get("updated", "")[:10] if fields.get("updated") else "",
                })

            total = data.get("total", len(issues))
            return {
                "issues": issues,
                "result_type": "issues",
                "response": f"Found {total} issue(s)",
                "jira_results": issues,
            }
        except Exception as direct_err:
            logger.warning(f"[FAST] direct jira_search_issues failed, trying MCP fallback: {direct_err}")
            result = _mcp_call("jira_search_issues", {"jql": jql, "max_results": 50})
            if result.get("error"):
                return {
                    "error": f"Issue lookup failed: {direct_err} | MCP fallback failed: {result.get('error')}",
                    "response": "Could not fetch issues. Please retry.",
                }
            issues = result.get("issues", [])
            return {
                "issues": issues,
                "result_type": "issues",
                "response": f"Found {result.get('total', len(issues))} issue(s)",
                "jira_results": issues,
            }

    if any(p in prompt_lower for p in ["who is working", "who has", "assignee breakdown"]):
        # Honor explicit project hints in prompts like "who is working on what in ITSUP".
        project_hint_match = re.search(r"\bin\s+([a-z][a-z0-9]+)\b", prompt_lower)
        effective_project_key = project_key
        if project_hint_match:
            hinted = project_hint_match.group(1).upper().strip()
            if 1 <= len(hinted) <= 12:
                effective_project_key = hinted

        logger.info(f"[FAST] jira_get_project_summary (assignees) for {effective_project_key}")
        result = _mcp_call("jira_get_project_summary", {"project_key": effective_project_key})
        if result.get("error"):
            return {"error": result["error"], "response": result["error"]}

        counts = result.get("assignee_counts", {})
        if not counts:
            # Fallback: aggregate assignees directly from recent project issues.
            try:
                data = jira_get_cached("search/jql", params={
                    "jql": f"project={effective_project_key} ORDER BY created DESC",
                    "maxResults": 100,
                    "fields": "assignee",
                })
                for issue in data.get("issues", []):
                    assignee = issue.get("fields", {}).get("assignee")
                    name = assignee.get("displayName", "Unassigned") if assignee else "Unassigned"
                    counts[name] = counts.get(name, 0) + 1
            except Exception as e:
                logger.warning(f"[FAST] assignee fallback aggregation failed: {e}")

        results = [{"name": k, "count": v} for k, v in counts.items()]
        return {"result_type": "assignees", "jira_results": results,
                "response": f"Found {len(results)} assignees"}

    # Burndown/outlook narrative with actionable To Do reasons.
    if ("burndown" in prompt_lower or "burn down" in prompt_lower or "outlook" in prompt_lower) and ("to do" in prompt_lower):
        # Honor explicit project hints in prompts such as "for ITSUP" or "in ITSUP".
        effective_project_key = project_key
        key_hint_match = re.search(r"\b(?:for|in)\s+([a-z][a-z0-9]+)\b", prompt_lower)
        if key_hint_match:
            hinted = key_hint_match.group(1).upper().strip()
            if 1 <= len(hinted) <= 12:
                effective_project_key = hinted

        try:
            data = jira_get_cached("search/jql", params={
                "jql": f"project={effective_project_key} ORDER BY updated DESC",
                "maxResults": 200,
                "fields": "summary,status,assignee,priority,created,updated,issuetype,labels",
            })
            raw_issues = data.get("issues", [])

            def _map_status(status_name: str) -> str:
                s = (status_name or "").strip().lower()
                if s in {"to do"}:
                    return "To Do"
                if s in {"in progress", "in review", "review", "testing", "quality assurance", "qa"}:
                    return "In Progress"
                if s in {"done", "closed", "resolved", "complete", "completed"}:
                    return "Done"
                if s in {"blocked", "impediment"}:
                    return "Blocked"
                return "Other"

            issues = []
            todo_issues = []
            now = datetime.now(timezone.utc)
            stale_todo = 0
            unassigned_todo = 0
            high_priority_todo = 0
            blocker_marked_todo = 0

            for item in raw_issues:
                fields = item.get("fields", {})
                status_name = fields.get("status", {}).get("name", "") if fields.get("status") else ""
                mapped_status = _map_status(status_name)
                assignee = fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned"
                priority = fields.get("priority", {}).get("name", "") if fields.get("priority") else ""
                issue = {
                    "key": item.get("key", ""),
                    "summary": fields.get("summary", ""),
                    "status": status_name,
                    "assignee": assignee,
                    "priority": priority,
                    "type": fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else "",
                    "created": fields.get("created", "")[:10] if fields.get("created") else "",
                    "updated": fields.get("updated", "")[:10] if fields.get("updated") else "",
                }
                issues.append(issue)

                if mapped_status == "To Do":
                    todo_issues.append(issue)
                    if assignee == "Unassigned":
                        unassigned_todo += 1
                    if priority.lower() in {"high", "highest", "critical", "blocker"}:
                        high_priority_todo += 1
                    labels = [str(lbl).lower() for lbl in fields.get("labels", [])]
                    summary = (fields.get("summary") or "").lower()
                    if any(term in labels for term in ["blocked", "blocker", "impediment"]) or any(term in summary for term in ["blocked", "blocker", "impediment"]):
                        blocker_marked_todo += 1
                    updated_at = parse_jira_datetime(fields.get("updated", ""))
                    if updated_at and (now - updated_at) > timedelta(days=7):
                        stale_todo += 1

            recent_done = jira_get_cached("search/jql", params={
                "jql": f"project={effective_project_key} AND statusCategory = Done AND resolved >= -14d",
                "maxResults": 1,
                "fields": "resolutiondate",
            }).get("total", 0)
            recent_created = jira_get_cached("search/jql", params={
                "jql": f"project={effective_project_key} AND created >= -14d",
                "maxResults": 1,
                "fields": "created",
            }).get("total", 0)

            completion_rate = round((recent_done / max(recent_created, 1)) * 100, 1) if recent_created else 100.0
            net_flow = recent_done - recent_created
            if net_flow >= 0:
                trend = f"Burn-down trend is stable to improving ({recent_done} done vs {recent_created} created in 14d)."
            else:
                trend = f"Burn-down risk: intake exceeds completion ({recent_done} done vs {recent_created} created in 14d)."

            blockers = []
            if unassigned_todo:
                blockers.append(f"{unassigned_todo} unassigned")
            if stale_todo:
                blockers.append(f"{stale_todo} stale (>7d)")
            if high_priority_todo:
                blockers.append(f"{high_priority_todo} high-priority")
            if blocker_marked_todo:
                blockers.append(f"{blocker_marked_todo} blocker-tagged")
            reasons = ", ".join(blockers) if blockers else "no major blockers detected"

            response = (
                f"{trend} To Do has {len(todo_issues)} issue(s); key hold factors: {reasons}. "
                f"Completion ratio is {completion_rate}% over the last 14 days."
            )

            return {
                "result_type": "issues",
                "jira_results": todo_issues[:50],
                "response": response,
            }
        except Exception as e:
            logger.warning(f"[FAST] burndown outlook failed: {e}")
            return {
                "response": f"Burndown analysis failed: {e}",
            }

    return None


# ══════════════════════════════════════════════════════════════════════════════
# MCP Agent
# ══════════════════════════════════════════════════════════════════════════════

def run_jira_agent(prompt: str, project_key: str) -> Dict:
    if not llm:
        return {"error": "LLM not configured", "raw_response": ""}

    auth_context = get_mcp_auth_context()
    if not auth_context:
        return {"error": "Jira auth context is missing", "raw_response": ""}

    try:
        tool_specs = list_jira_tools()
    except Exception as e:
        logger.error(f"MCP tools/list failed: {e}")
        return {"error": f"MCP unavailable: {e}", "raw_response": ""}

    tool_map = {t.get("name"): t for t in tool_specs if t.get("name")}
    tool_descriptions = "\n".join([
        f"- {t.get('name')}: {(t.get('description') or '').strip().split(chr(10))[0]}"
        for t in tool_specs
    ])

    system_prompt = f"""You are a Jira assistant. Current project key: {project_key}

Available tools:
{tool_descriptions}

To use a tool respond ONLY with JSON:
{{"tool": "tool_name", "args": {{"arg1": "value1"}}}}

To give final answer:
{{"result": {{...}}, "summary": "explanation"}}

Always use project key: {project_key}
"""
    messages = [SystemMessage(content=system_prompt), HumanMessage(content=prompt)]
    tool_results = []

    def _extract_first_json_object(text: str) -> Optional[str]:
        if not text:
            return None
        start = text.find("{")
        if start < 0:
            return None

        depth = 0
        in_string = False
        escape = False
        for idx in range(start, len(text)):
            ch = text[idx]
            if in_string:
                if escape:
                    escape = False
                elif ch == "\\":
                    escape = True
                elif ch == '"':
                    in_string = False
                continue

            if ch == '"':
                in_string = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return text[start:idx + 1]
        return None

    def _safe_json_loads(blob: str) -> Optional[Dict[str, Any]]:
        if not blob:
            return None
        try:
            parsed = json.loads(blob)
            return parsed if isinstance(parsed, dict) else None
        except json.JSONDecodeError:
            pass

        # Common LLM JSON mistakes: trailing commas and control characters.
        cleaned = re.sub(r",\s*([}\]])", r"\1", blob)
        cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", cleaned)
        try:
            parsed = json.loads(cleaned)
            return parsed if isinstance(parsed, dict) else None
        except json.JSONDecodeError:
            return None

    for i in range(5):
        try:
            response = llm.invoke(messages)
            raw_text = response.content.strip()
            logger.info(f"Agent iter {i+1}: {raw_text[:200]}")

            # OpenRouter responses can include extra trailing text; decode first valid JSON object only.
            json_blob = _extract_first_json_object(raw_text)
            if not json_blob:
                return {"raw_response": raw_text}

            parsed = _safe_json_loads(json_blob)
            if parsed is None:
                logger.warning(f"Agent iter {i+1}: invalid JSON payload; requesting retry")
                messages.append(response)
                messages.append(HumanMessage(
                    content=(
                        "Your previous message was not valid JSON. "
                        "Return ONLY valid JSON with one of these exact shapes: "
                        "{\"tool\":\"name\",\"args\":{...}} or {\"result\":{...},\"summary\":\"...\"}. "
                        "Do not include markdown fences or extra text."
                    )
                ))
                continue

            if "tool" in parsed:
                tool_name = parsed.get("tool")
                tool_args = parsed.get("args", {})
                if tool_name not in tool_map:
                    return {"raw_response": raw_text}
                tool_args = tool_args or {}
                tool_args["auth_context"] = auth_context
                if tool_name in {"jira_get_sprints", "jira_get_project_summary", "jira_get_my_issues", "jira_create_issue"}:
                    tool_args.setdefault("project_key", project_key)
                logger.info(f"MCP tool call: {tool_name}({tool_args})")
                try:
                    tool_result = call_jira_tool(tool_name, tool_args)
                    tool_results.append({"tool": tool_name, "result": tool_result})
                    messages.append(response)
                    messages.append(HumanMessage(
                        content=f"Tool {tool_name} returned: {json.dumps(tool_result)}\n\nNow give the final result as JSON."
                    ))
                except Exception as e:
                    if tool_name == "jira_search_issues":
                        try:
                            jql = tool_args.get("jql", f"project={project_key} ORDER BY created DESC")
                            max_results = int(tool_args.get("max_results", 50) or 50)
                            data = jira_get_cached("search/jql", params={
                                "jql": jql,
                                "maxResults": max_results,
                                "fields": "summary,status,assignee,priority,created,updated,issuetype",
                            })
                            raw_issues = data.get("issues", [])
                            issues = []
                            for item in raw_issues:
                                fields = item.get("fields", {})
                                issues.append({
                                    "key": item.get("key", ""),
                                    "summary": fields.get("summary", ""),
                                    "status": fields.get("status", {}).get("name", ""),
                                    "assignee": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
                                    "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
                                    "type": fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else "",
                                    "created": fields.get("created", "")[:10] if fields.get("created") else "",
                                    "updated": fields.get("updated", "")[:10] if fields.get("updated") else "",
                                })
                            tool_result = {"issues": issues, "total": data.get("total", len(issues))}
                            tool_results.append({"tool": "jira_search_issues_rest_fallback", "result": tool_result})
                            messages.append(response)
                            messages.append(HumanMessage(
                                content=f"Tool jira_search_issues (REST fallback) returned: {json.dumps(tool_result)}\n\nNow give the final result as JSON."
                            ))
                            continue
                        except Exception as fallback_err:
                            logger.warning(f"Agent REST fallback failed for jira_search_issues: {fallback_err}")
                    return {"raw_response": f"Tool {tool_name} failed: {e}"}

            elif "result" in parsed:
                result = parsed["result"]
                result["raw_response"] = parsed.get("summary", raw_text)
                result["tool_calls"]   = [t["tool"] for t in tool_results]
                return result
            else:
                return {"raw_response": raw_text}

        except Exception as e:
            logger.error(f"Agent iter {i+1} error: {e}")
            messages.append(HumanMessage(
                content=(
                    "An internal error occurred while parsing your response. "
                    "Please return strict JSON only in one object."
                )
            ))
            continue

    if tool_results:
        last = tool_results[-1]["result"]
        last["raw_response"] = f"Completed via {[t['tool'] for t in tool_results]}"
        return last
    return {"raw_response": "Could not complete the request"}


# ══════════════════════════════════════════════════════════════════════════════
# Chart generation
# ══════════════════════════════════════════════════════════════════════════════

def generate_chart_b64(chart_type: str, title: str, labels: list, values: list) -> str:
    palette = ["#0f9d8a", "#2f80ed", "#ff6b4a", "#f0b429", "#5b6cff", "#2eaf74"]
    labels = labels or ["No Data"]
    raw_values = values or [0]
    numeric_values = []
    for v in raw_values:
        try:
            numeric_values.append(float(v))
        except Exception:
            numeric_values.append(0.0)

    fig, ax = plt.subplots(figsize=(10, 6), dpi=150)
    fig.patch.set_facecolor("#f8fafc")
    ax.set_facecolor("#ffffff")

    colors = [palette[i % len(palette)] for i in range(max(len(labels), 1))]
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#d9e2ef")
    ax.spines["bottom"].set_color("#d9e2ef")

    if chart_type == "pie":
        total = sum(numeric_values)
        pie_values = numeric_values if total > 0 else [1]
        pie_labels = labels if total > 0 else ["No Data"]
        ax.pie(
            pie_values,
            labels=pie_labels,
            autopct="%1.1f%%" if total > 0 else None,
            colors=colors[:len(pie_labels)],
            startangle=90,
            wedgeprops={"width": 0.55, "edgecolor": "white", "linewidth": 2},
            textprops={"fontsize": 10, "color": "#24344d"},
        )
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=16)

    elif chart_type == "line":
        ax.plot(labels, numeric_values, marker="o", linewidth=3, markersize=7, color=palette[1])
        ax.fill_between(range(len(labels)), numeric_values, color=palette[1], alpha=0.14)
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=12)
        ax.set_xlabel("Categories", color="#66758f", fontsize=11)
        ax.set_ylabel("Values", color="#66758f", fontsize=11)
        ax.tick_params(axis="x", rotation=28, labelsize=10)
        ax.tick_params(axis="y", labelsize=10)
        ax.grid(axis="y", alpha=0.25, linestyle="--")

    elif chart_type == "metrics":
        ax.axis("off")
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=10)
        for i, (label, value) in enumerate(zip(labels, raw_values)):
            ax.text(
                0.5,
                0.84 - i * 0.18,
                f"{label}: {value}",
                ha="center",
                va="center",
                fontsize=15,
                fontweight="bold",
                color="#1f2f46",
                bbox={
                    "boxstyle": "round,pad=0.45",
                    "facecolor": colors[i % len(colors)],
                    "edgecolor": "none",
                    "alpha": 0.18,
                },
                transform=ax.transAxes,
            )

    elif chart_type == "gantt":
        timeline_rows = []
        for idx, label in enumerate(labels):
            timeline_rows.append({
                "label": str(label),
                "start": None,
                "end": None,
                "duration": numeric_values[idx] if idx < len(numeric_values) else 0.0,
                "status": "",
            })

        if isinstance(values, list) and values and isinstance(values[0], dict):
            timeline_rows = []
            for row in values:
                if not isinstance(row, dict):
                    continue
                timeline_rows.append({
                    "label": str(row.get("label") or row.get("key") or "Issue"),
                    "start": row.get("start"),
                    "end": row.get("end"),
                    "duration": row.get("duration", 1),
                    "status": str(row.get("status", "")),
                })

        done_color = "#2eaf74"
        in_progress_color = "#2f80ed"
        todo_color = "#f0b429"
        other_color = "#9aa8bd"

        def _status_color(status_name: str) -> str:
            s = (status_name or "").strip().lower()
            if s in {"done", "closed", "resolved", "complete", "completed"}:
                return done_color
            if s in {"in progress", "in review", "review", "testing", "qa"}:
                return in_progress_color
            if s in {"to do", "todo", "open", "backlog"}:
                return todo_color
            return other_color

        parsed_rows = []
        fallback_base = datetime.now(timezone.utc)
        for row in timeline_rows:
            start_dt = parse_jira_datetime(str(row.get("start") or ""))
            end_dt = parse_jira_datetime(str(row.get("end") or ""))
            if not start_dt and end_dt:
                start_dt = end_dt - timedelta(days=max(float(row.get("duration", 1) or 1), 1.0))
            if not end_dt and start_dt:
                end_dt = start_dt + timedelta(days=max(float(row.get("duration", 1) or 1), 1.0))
            if not start_dt and not end_dt:
                start_dt = fallback_base + timedelta(days=len(parsed_rows))
                end_dt = start_dt + timedelta(days=max(float(row.get("duration", 1) or 1), 1.0))
            if end_dt < start_dt:
                end_dt = start_dt + timedelta(days=1)

            parsed_rows.append({
                "label": row.get("label", "Issue"),
                "start": start_dt,
                "end": end_dt,
                "status": row.get("status", ""),
            })

        parsed_rows = parsed_rows[:20]
        y_positions = list(range(len(parsed_rows)))
        bar_labels = [row["label"] for row in parsed_rows]
        left_edges = [mdates.date2num(row["start"]) for row in parsed_rows]
        widths = [
            max(mdates.date2num(row["end"]) - mdates.date2num(row["start"]), 0.8)
            for row in parsed_rows
        ]
        colors = [_status_color(row["status"]) for row in parsed_rows]

        ax.barh(y_positions, widths, left=left_edges, color=colors, edgecolor="none", alpha=0.95)
        ax.set_yticks(y_positions)
        ax.set_yticklabels(bar_labels, fontsize=9, color="#24344d")
        ax.invert_yaxis()
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=12)
        ax.set_xlabel("Timeline", color="#66758f", fontsize=11)
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b"))
        ax.tick_params(axis="x", labelsize=9, rotation=25)
        ax.grid(axis="x", alpha=0.25, linestyle="--")

    else:
        bars = ax.bar(labels, numeric_values, color=colors[:len(labels)], edgecolor="none")
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=12)
        ax.set_xlabel("Categories", color="#66758f", fontsize=11)
        ax.set_ylabel("Values", color="#66758f", fontsize=11)
        ax.tick_params(axis="x", rotation=28, labelsize=10)
        ax.tick_params(axis="y", labelsize=10)
        ax.grid(axis="y", alpha=0.25, linestyle="--")
        y_max = max(numeric_values) if numeric_values else 0
        pad = max(y_max * 0.02, 0.5)
        for bar, value in zip(bars, raw_values):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + pad,
                str(value),
                ha="center",
                va="bottom",
                fontsize=10,
                color="#24344d",
            )

    plt.tight_layout(pad=1.1)
    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(); buf.seek(0)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


def detect_chart_type(prompt: str) -> str:
    p = prompt.lower()
    if any(w in p for w in ["gantt", "timeline", "roadmap"]): return "gantt"
    if "pie"   in p: return "pie"
    if "line"  in p: return "line"
    if any(w in p for w in ["metric", "kpi"]): return "metrics"
    return "bar"


def build_gantt_chart_data(prompt: str, project_key: str) -> Optional[Dict[str, Any]]:
    effective_project_key = project_key
    key_hint_match = re.search(r"\b(?:for|in)\s+([a-z][a-z0-9]+)\b", prompt.lower())
    if key_hint_match:
        hinted = key_hint_match.group(1).upper().strip()
        if 1 <= len(hinted) <= 12:
            effective_project_key = hinted

    try:
        data = jira_get_cached("search/jql", params={
            "jql": f"project={effective_project_key} ORDER BY duedate ASC, created ASC",
            "maxResults": 40,
            "fields": "summary,status,created,updated,duedate,resolutiondate,priority",
        })
    except Exception as e:
        logger.warning(f"Gantt data fetch failed: {e}")
        return None

    rows = []
    for issue in data.get("issues", []):
        fields = issue.get("fields", {})
        key = issue.get("key", "")
        status_name = fields.get("status", {}).get("name", "") if fields.get("status") else ""
        created_dt = parse_jira_datetime(fields.get("created", ""))
        due_dt = parse_jira_datetime(fields.get("duedate", ""))
        resolved_dt = parse_jira_datetime(fields.get("resolutiondate", ""))

        start_dt = created_dt or due_dt or resolved_dt
        if not start_dt:
            continue

        if is_done_status(status_name) and resolved_dt:
            end_dt = resolved_dt
        elif due_dt:
            end_dt = due_dt
        else:
            end_dt = start_dt + timedelta(days=14)

        if end_dt < start_dt:
            end_dt = start_dt + timedelta(days=1)

        duration_days = max((end_dt - start_dt).days, 1)
        rows.append({
            "label": key,
            "start": start_dt.isoformat(),
            "end": end_dt.isoformat(),
            "duration": duration_days,
            "status": status_name,
            "summary": fields.get("summary", ""),
            "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
        })

    rows.sort(key=lambda r: r.get("start", ""))
    rows = rows[:20]

    if not rows:
        return None

    return {
        "title": f"{effective_project_key} Delivery Timeline (Gantt)",
        "labels": [r["label"] for r in rows],
        "values": rows,
    }


def extract_chart_data(agent_result: dict, prompt: str, project_key: str) -> Optional[Dict]:
    if "chart_data" in agent_result:
        cd = agent_result["chart_data"]
        return {"title": cd.get("title", "Chart"), "labels": cd.get("labels", []), "values": cd.get("values", [])}

    for key, title_template in [
        ("status_counts",   f"{project_key} — Issue Status"),
        ("assignee_counts", f"{project_key} — Issues by Assignee"),
        ("priority_counts", f"{project_key} — Issues by Priority"),
        ("type_counts",     f"{project_key} — Issues by Type"),
    ]:
        if key in agent_result and agent_result[key]:
            p = prompt.lower()
            selected_key = key
            if "assignee" in p or "who" in p:     selected_key = "assignee_counts"
            elif "priority" in p:                  selected_key = "priority_counts"
            elif "type" in p or "kind" in p:       selected_key = "type_counts"
            else:                                  selected_key = "status_counts"
            counts = agent_result.get(selected_key, agent_result[key])
            return {"title": title_template, "labels": list(counts.keys()), "values": list(counts.values())}
    return None


def ensure_jira_chart_data_for_export(prompt: str, project_key: str, agent_result: dict) -> Optional[Dict]:
    """
    For exports, ensure we fetch REAL Jira data.
    Tries multiple strategies to get real data instead of dummy data.
    """
    if not project_key:
        return None

    if detect_chart_type(prompt) == "gantt":
        gantt_chart = build_gantt_chart_data(prompt, project_key)
        if gantt_chart:
            return gantt_chart
    
    # Try to extract from agent result first
    chart_data = extract_chart_data(agent_result, prompt, project_key)
    if chart_data:
        logger.info(f"[EXPORT] Using Jira data from agent result")
        return chart_data
    
    # If not found, call project summary directly
    try:
        auth_context = get_mcp_auth_context()
        if auth_context:
            summary = call_jira_tool(
                "jira_get_project_summary",
                {
                    "auth_context": auth_context,
                    "project_key": project_key,
                },
            )
            logger.info(f"[EXPORT] Fetched project summary for {project_key}")
            chart_data = extract_chart_data(summary, prompt, project_key)
            if chart_data:
                logger.info(f"[EXPORT] Using Jira data from project summary")
                return chart_data
    except Exception as e:
        logger.warning(f"[EXPORT] Failed to fetch project summary: {e}")
    
    # Alternative: search for issues and aggregate by status
    try:
        if get_jira_auth():
            data = jira_get_cached("search/jql", params={
                "jql": f"project={project_key} ORDER BY created DESC",
                "maxResults": 100,
                "fields": "status,priority,assignee,issuetype",
            })
            logger.info(f"[EXPORT] Fetched {data.get('total', 0)} issues from Jira")
            
            # Map Jira status names to fixed dashboard categories before counting
            def map_status(status_name):
                s = (status_name or "").strip().lower()
                if s in {"to do"}:
                    return "To Do"
                elif s in {"in progress", "in review", "review", "testing"}:
                    return "In Progress"
                elif s in {"done", "closed", "resolved", "complete", "completed"}:
                    return "Done"
                elif s in {"blocked", "impediment"}:
                    return "Blocked"
                elif s in {"quality assurance", "qa"}:
                    return "Quality Assurance"
                else:
                    return "Other"

            status_counts = {"To Do": 0, "In Progress": 0, "Done": 0, "Blocked": 0, "Quality Assurance": 0, "Other": 0}
            for issue in data.get("issues", []):
                raw_status = issue.get("fields", {}).get("status", {}).get("name", "Unknown")
                mapped_status = map_status(raw_status)
                status_counts[mapped_status] += 1

            # Remove zero-count categories except 'Other' if it's zero
            filtered_counts = {k: v for k, v in status_counts.items() if v > 0 and (k != "Other" or v > 0)}
            if filtered_counts:
                logger.info(f"[EXPORT] Using aggregated status data: {filtered_counts}")
                return {
                    "title": f"{project_key} — Issue Status",
                    "labels": list(filtered_counts.keys()),
                    "values": list(filtered_counts.values())
                }
    except Exception as e:
        logger.warning(f"[EXPORT] Failed to aggregate issues by status: {e}")
    
    logger.warning(f"[EXPORT] No real Jira data found for {project_key}")
    return None


# ══════════════════════════════════════════════════════════════════════════════
# Excel Export with Chart
# ══════════════════════════════════════════════════════════════════════════════

def create_excel_with_chart(chart_data: dict, chart_config: dict, project_key: str = "") -> BytesIO:
    """
    Create an Excel file with:
    - Sheet 1: Aggregated data from chart (status counts, etc.)
    - Sheet 2: Detailed Jira issues
    - Sheet 3: Chart image embedded
    """
    from openpyxl.styles import Font, Alignment
    
    # Create workbook
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Add summary headers
    ws_summary['A1'] = "Category"
    ws_summary['B1'] = "Count"
    ws_summary['A1'].font = Font(bold=True, size=12)
    ws_summary['B1'].font = Font(bold=True, size=12)
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary['B1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add summary data rows
    labels = chart_data.get("labels", [])
    values = chart_data.get("values", [])
    for idx, (label, value) in enumerate(zip(labels, values), start=2):
        ws_summary[f'A{idx}'] = label
        ws_summary[f'B{idx}'] = value
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    # Add detailed Jira issues sheet
    ws_issues = wb.create_sheet("Jira Issues")
    ws_issues['A1'] = "Issue Key"
    ws_issues['B1'] = "Summary"
    ws_issues['C1'] = "Status"
    ws_issues['D1'] = "Assignee"
    ws_issues['E1'] = "Priority"
    ws_issues['F1'] = "Type"
    ws_issues['G1'] = "Created"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_issues[f'{col}1'].font = Font(bold=True, size=11)
        ws_issues[f'{col}1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Fetch detailed issues from Jira
    try:
        if project_key and get_jira_auth():
            data = jira_get_cached("search/jql", params={
                "jql": f"project={project_key} ORDER BY created DESC",
                "maxResults": 50,
                "fields": "key,summary,status,assignee,priority,issuetype,created",
            })
            
            row = 2
            for issue in data.get("issues", []):
                fields = issue.get("fields", {})
                ws_issues[f'A{row}'] = issue.get("key", "")
                ws_issues[f'B{row}'] = fields.get("summary", "")
                ws_issues[f'C{row}'] = fields.get("status", {}).get("name", "")
                ws_issues[f'D{row}'] = fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned"
                ws_issues[f'E{row}'] = fields.get("priority", {}).get("name", "") if fields.get("priority") else ""
                ws_issues[f'F{row}'] = fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else ""
                ws_issues[f'G{row}'] = fields.get("created", "")[:10] if fields.get("created") else ""
                row += 1
            
            logger.info(f"[EXCEL] Added {row - 2} detailed Jira issues for {project_key}")
    except Exception as e:
        logger.warning(f"Could not fetch detailed Jira issues for {project_key}: {e}")
    
    # Adjust column widths for issues sheet
    ws_issues.column_dimensions['A'].width = 12
    ws_issues.column_dimensions['B'].width = 35
    ws_issues.column_dimensions['C'].width = 15
    ws_issues.column_dimensions['D'].width = 20
    ws_issues.column_dimensions['E'].width = 12
    ws_issues.column_dimensions['F'].width = 12
    ws_issues.column_dimensions['G'].width = 12
    
    # Add chart image to third sheet
    if chart_config.get("image_base64"):
        ws_chart = wb.create_sheet("Chart")
        
        try:
            # Decode base64 image
            img_b64 = chart_config["image_base64"]
            if img_b64.startswith("data:image/png;base64,"):
                img_b64 = img_b64.replace("data:image/png;base64,", "")
            
            img_data = BytesIO(base64.b64decode(img_b64))
            img_data.seek(0)
            
            # Add image to worksheet
            img = XLImage(img_data)
            img.width = 600
            img.height = 400
            ws_chart.add_image(img, "A1")
            logger.info(f"Chart image embedded in Excel export")
        except Exception as e:
            logger.warning(f"Could not embed chart image in Excel: {e}")
    
    # Save to BytesIO
    excel_buf = BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)
    return excel_buf


# ══════════════════════════════════════════════════════════════════════════════
# Main process function
# ══════════════════════════════════════════════════════════════════════════════
def _top_counts(items: List[Dict[str, Any]], key_name: str, fallback: str = "Unknown", limit: int = 6) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for item in items:
        raw = item.get(key_name)
        label = str(raw).strip() if raw not in (None, "") else fallback
        counts[label] = counts.get(label, 0) + 1

    ranked = sorted(counts.items(), key=lambda pair: pair[1], reverse=True)
    top = ranked[:limit]
    remainder = ranked[limit:]
    if remainder:
        top.append(("Other", sum(value for _, value in remainder)))
    return dict(top)


def build_multi_chart_dashboard(project_key: str) -> Dict[str, Any]:
    if not get_jira_auth():
        raise ValueError("Jira not connected")

    data = jira_get_cached("search/jql", params={
        "jql": f"project={project_key} ORDER BY created DESC",
        "maxResults": 200,
        "fields": "summary,status,assignee,priority,issuetype,created,updated",
    })

    issues = []
    for issue in data.get("issues", []):
        fields = issue.get("fields", {})
        issues.append({
            "status": fields.get("status", {}).get("name", "Unknown") if fields.get("status") else "Unknown",
            "assignee": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
            "priority": fields.get("priority", {}).get("name", "None") if fields.get("priority") else "None",
            "type": fields.get("issuetype", {}).get("name", "Unknown") if fields.get("issuetype") else "Unknown",
        })

    widgets_spec = [
        {
            "title": f"{project_key} Issue Status",
            "chart_type": "bar",
            "counts": _top_counts(issues, "status", "Unknown", 8),
        },
        {
            "title": f"{project_key} Issues by Assignee",
            "chart_type": "bar",
            "counts": _top_counts(issues, "assignee", "Unassigned", 6),
        },
        {
            "title": f"{project_key} Issues by Priority",
            "chart_type": "pie",
            "counts": _top_counts(issues, "priority", "None", 6),
        },
        {
            "title": f"{project_key} Issues by Type",
            "chart_type": "pie",
            "counts": _top_counts(issues, "type", "Unknown", 6),
        },
    ]

    widgets = []
    for spec in widgets_spec:
        labels = list(spec["counts"].keys())
        values = list(spec["counts"].values())
        widgets.append({
            "chart_id": str(uuid.uuid4())[:8],
            "chart_type": spec["chart_type"],
            "title": spec["title"],
            "labels": labels,
            "values": values,
            "image_base64": generate_chart_b64(spec["chart_type"], spec["title"], labels, values),
            "data_source": "jira_mcp",
            "created_at": datetime.now().isoformat(),
        })

    return {
        "dashboard_id": str(uuid.uuid4())[:8],
        "title": f"{project_key} Multi-Chart Dashboard",
        "widgets": widgets,
        "created_at": datetime.now().isoformat(),
        "data_source": "jira_mcp",
    }

def process_prompt(prompt: str) -> Dict:
    prompt_lower = prompt.lower().strip()
    project_key  = get_user_project_key()

    def _format_risk_matrix_result(payload: Dict[str, Any], fallback_summary: str = "") -> Optional[Dict[str, Any]]:
        matrix = payload.get("risk_matrix")
        if not isinstance(matrix, dict):
            matrix = payload.get("riskMatrix")

        if not isinstance(matrix, dict):
            return None

        # Shape: {"riskMatrix": {"rows": [{"priority":"High","issues":[...]}]}}
        if isinstance(matrix.get("rows"), list):
            flattened = []
            for row in matrix.get("rows", []):
                if not isinstance(row, dict):
                    continue
                priority = row.get("priority", "")
                for item in row.get("issues", []) or []:
                    if not isinstance(item, dict):
                        continue
                    flattened.append({
                        "key": item.get("key", ""),
                        "summary": item.get("summary", ""),
                        "status": item.get("status", ""),
                        "priority": priority,
                        "assignee": item.get("assignee", "Unassigned"),
                    })

            summary = payload.get("summary") or fallback_summary or "Risk matrix generated."
            return {
                "response": summary,
                "result_type": "issues",
                "jira_results": flattened[:80],
            }

        flattened = []
        for priority, status_map in matrix.items():
            if not isinstance(status_map, dict):
                continue
            for status, items in status_map.items():
                if not isinstance(items, list):
                    continue
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    flattened.append({
                        "key": item.get("key", ""),
                        "summary": item.get("summary", ""),
                        "status": status,
                        "priority": priority,
                        "assignee": item.get("assignee", "Unassigned"),
                    })

        summary = payload.get("summary") or fallback_summary or "Risk matrix generated."
        return {
            "response": summary,
            "result_type": "issues",
            "jira_results": flattened[:80],
        }

    if not project_key:
        return {"error": "No project selected. Please select a Jira project first.", "needs_project": True}

    # ── Special: clear all dashboards ────────────────────────────────────────
    if any(w in prompt_lower for w in ["clear all", "clear dashboard", "remove all", "delete all"]):
        uid = user_id()
        if uid and uid in _user_dashboards:
            _user_dashboards[uid] = {}
        return {"response": "All dashboards cleared! ✓", "cleared": True}

    chart_keywords  = ["chart", "graph", "pie", "bar", "line", "visuali", "plot", "metrics", "kpi", "burndown", "burn down", "gantt", "timeline", "roadmap"]
    create_keywords = ["create task", "create a task", "new task", "add task",
                       "create ticket", "open a ticket", "raise a ticket"]
    export_keywords = ["export", "excel", "download", "xlsx", "csv"]

    wants_chart  = any(w in prompt_lower for w in chart_keywords)
    wants_create = any(p in prompt_lower for p in create_keywords)
    wants_export = any(w in prompt_lower for w in export_keywords)
    chart_type   = detect_chart_type(prompt)
    jira_ok      = get_jira_auth() is not None
    wants_multi_chart_dashboard = any(
        phrase in prompt_lower
        for phrase in [
            "dashboard",
            "multiple charts",
            "multi chart",
            "overview charts",
            "chart dashboard",
            "multiple chart dashboard",
            "multi-chart dashboard",
        ]
    )

    if wants_chart:
        if wants_multi_chart_dashboard:
            try:
                dashboard_config = build_multi_chart_dashboard(project_key)
                return {
                    "response": "Multi-chart dashboard generated.",
                    "dashboard_config": dashboard_config,
                }
            except Exception as e:
                logger.warning(f"Multi-chart dashboard generation failed: {e}")
    
    # Fast dispatch for non-chart queries
    if jira_ok and not wants_chart and not wants_create:
        fast = fast_dispatch(prompt_lower, project_key)
        if fast:
            logger.info("[FAST DISPATCH] LLM skipped")
            return fast

    # Run MCP agent
    agent_result = {}
    if jira_ok:
        agent_result = run_jira_agent(prompt, project_key)

    # Chart path
    if wants_chart:
        if wants_multi_chart_dashboard:
            try:
                dashboard_config = build_multi_chart_dashboard(project_key)
                return {
                    "response": "Multi-chart dashboard generated.",
                    "dashboard_config": dashboard_config,
                }
            except Exception as e:
                logger.warning(f"Multi-chart dashboard generation failed: {e}")

        # For exports, prioritize real Jira data
        chart_data_source = "generated"
        chart_data = None

        if chart_type == "gantt" and jira_ok:
            chart_data = build_gantt_chart_data(prompt, project_key)
            if chart_data:
                chart_data_source = "jira_mcp"

        if wants_export:
            chart_data = chart_data or ensure_jira_chart_data_for_export(prompt, project_key, agent_result)
            if chart_data:
                chart_data_source = "jira_mcp"
        else:
            chart_data = chart_data or extract_chart_data(agent_result, prompt, project_key)
            if chart_data and jira_ok:
                chart_data_source = "jira_mcp"

        if not chart_data and jira_ok:
            try:
                summary = call_jira_tool(
                    "jira_get_project_summary",
                    {
                        "auth_context": get_mcp_auth_context() or {},
                        "project_key": project_key,
                    },
                )
                chart_data = extract_chart_data(summary, prompt, project_key)
                if chart_data:
                    chart_data_source = "jira_mcp"
            except Exception as e:
                logger.warning(f"Project summary fallback failed: {e}")

        if not chart_data and llm:
            try:
                system = SystemMessage(content='Generate chart data as JSON only:\n{"title":"...","labels":[...],"values":[...]}')
                resp   = llm.invoke([system, HumanMessage(content=prompt)])
                match  = re.search(r"\{.*\}", resp.content, re.DOTALL)
                if match:
                    chart_data = json.loads(match.group())
                    chart_data_source = "llm_generated"
            except Exception:
                pass

        if not chart_data:
            logger.warning(f"[CHART] Using dummy data - no real Jira data available")
            chart_data = {"title": "Sample Distribution",
                          "labels": ["A", "B", "C", "D"], "values": [35, 25, 25, 15]}
            chart_data_source = "sample_fallback"

        chart_id = str(uuid.uuid4())[:8]

        # Auto-switch: bar/line with only 1 data point looks wrong → use pie or metrics
        effective_type = chart_type
        if len(chart_data["labels"]) == 1 and chart_type in ("bar", "line"):
            effective_type = "metrics"
            logger.info(f"Auto-switched chart type to metrics (only 1 data point)")

        image_b64 = generate_chart_b64(
            effective_type,
            chart_data["title"],
            chart_data["labels"],
            chart_data["values"],
        )
        chart_type = effective_type
        chart_config = {
            "chart_id":     chart_id,
            "chart_type":   chart_type,
            "title":        chart_data["title"],
            "labels":       chart_data["labels"],
            "values":       chart_data["values"],
            "image_base64": image_b64,
            "created_at":   datetime.now().isoformat(),
            "data_source":  chart_data_source,
        }

        result = {"response": f"{chart_type.upper()} chart generated.", "chart_config": chart_config}

        if wants_export:
            logger.info(f"[EXPORT] Creating Excel with chart data: {len(chart_data.get('labels', []))} items")
            excel_buf = create_excel_with_chart(chart_data, chart_config, project_key)
            chart_config["export_format"] = "excel"
            chart_config["export_data"]   = base64.b64encode(excel_buf.getvalue()).decode()
            result["auto_export"]  = True
            result["dashboard_id"] = chart_id
            result["data_source"]  = chart_config.get("data_source", "generated")
            logger.info(f"[EXPORT] Excel file generated successfully (dashboard_id: {chart_id})")

        return result

    # Create task path
    if wants_create:
        jira_task = agent_result.get("key") or agent_result.get("task_id")
        if jira_task:
            cfg    = get_jira_auth()
            server = cfg["base_url"] if cfg else ""
            if cfg and cfg["auth_type"] == "oauth":
                site   = session.get("jira_site_name", "")
                server = f"https://{site}.atlassian.net" if site else server
            return {
                "response":  f"Jira task created: {jira_task}",
                "jira_task": {"task_id": jira_task, "url": f"{server}/browse/{jira_task}", "summary": prompt[:100]},
            }

    # Query / text path
    if agent_result.get("error"):
        return {"response": agent_result.get("error") or "Request failed."}

    if agent_result.get("result_type"):
        return agent_result

    formatted_risk = _format_risk_matrix_result(agent_result, agent_result.get("raw_response", ""))
    if formatted_risk:
        return formatted_risk

    issues = agent_result.get("issues", [])
    if issues:
        return {"response": f"Found {len(issues)} issue(s)", "result_type": "issues", "jira_results": issues}

    raw = agent_result.get("raw_response", "")
    if raw and raw.lstrip().startswith("{"):
        try:
            parsed_raw = json.loads(raw)
            if isinstance(parsed_raw, dict) and isinstance(parsed_raw.get("result"), dict):
                nested = parsed_raw.get("result") or {}
                nested_summary = parsed_raw.get("summary") or nested.get("summary") or ""
                nested_risk = _format_risk_matrix_result(nested, nested_summary)
                if nested_risk:
                    return nested_risk
                if nested_summary:
                    return {"response": nested_summary}
        except Exception:
            pass

    if raw:
        return {"response": raw}

    return {"response": "Request processed."}


# ══════════════════════════════════════════════════════════════════════════════
# Login / OAuth Routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/login")
def login_page():
    """Show login page to unauthenticated users."""
    if is_authenticated():
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/auth/login")
def auth_login():
    """Start Atlassian OAuth flow — manual state to avoid CSRF issues."""
    import secrets
    state        = secrets.token_urlsafe(32)
    redirect_uri = url_for("auth_callback", _external=True)

    # Store state in a separate simple cookie (more reliable than session for OAuth)
    # prompt parameter — only passed when explicitly requested (e.g. after logout)
    prompt_param = f"&prompt={request.args.get('prompt', '')}" if request.args.get('prompt') else ""

    response = redirect(
        f"https://auth.atlassian.com/authorize"
        f"?audience=api.atlassian.com"
        f"&client_id={os.getenv('JIRA_CLIENT_ID')}"
        f"&scope={requests.utils.quote(JIRA_OAUTH_SCOPES, safe='')}"
        f"&redirect_uri={requests.utils.quote(redirect_uri, safe='')}"
        f"&state={state}"
        f"&response_type=code"
        f"{prompt_param}"
    )
    response.set_cookie("oauth_state", state, max_age=600, httponly=True, samesite="Lax")
    return response


@app.route("/auth/callback")
def auth_callback():
    """Handle OAuth callback — exchange code for token manually."""
    try:
        # Verify state from cookie (not session — more reliable)
        stored_state   = request.cookies.get("oauth_state", "")
        returned_state = request.args.get("state", "")

        if not stored_state or stored_state != returned_state:
            logger.error(f"State mismatch: stored={stored_state[:10]}... returned={returned_state[:10]}...")
            return render_template("login.html",
                error="Login session expired. Please try again.")

        # Exchange code for token manually
        code         = request.args.get("code")
        redirect_uri = url_for("auth_callback", _external=True)

        token_resp = requests.post(
            "https://auth.atlassian.com/oauth/token",
            json={
                "grant_type":    "authorization_code",
                "client_id":     os.getenv("JIRA_CLIENT_ID"),
                "client_secret": os.getenv("JIRA_CLIENT_SECRET"),
                "code":          code,
                "redirect_uri":  redirect_uri,
            },
            headers={"Content-Type": "application/json"}
        )
        token_resp.raise_for_status()
        token        = token_resp.json()
        access_token = token.get("access_token")

        # Fetch all accessible Jira sites first (always works)
        sites = fetch_accessible_sites(access_token)

        if not sites:
            return render_template("login.html",
                error="No Jira sites found for your account. Make sure you have access to at least one Jira project.")

        # Fetch user info — try /me first, fall back to Jira myself endpoint
        u = {}
        try:
            user_resp = requests.get(
                "https://api.atlassian.com/me",
                headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
            )
            user_resp.raise_for_status()
            u = user_resp.json()
            logger.info(f"User info from /me: {u.get('email', 'unknown')}")
        except Exception as me_err:
            logger.warning(f"/me endpoint failed ({me_err}), trying Jira myself endpoint")
            try:
                # Try getting user info from Jira API itself
                cloud_id = sites[0]["id"]
                myself_resp = requests.get(
                    f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/api/3/myself",
                    headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
                )
                myself_resp.raise_for_status()
                jira_user = myself_resp.json()
                u = {
                    "name":    jira_user.get("displayName", ""),
                    "email":   jira_user.get("emailAddress", ""),
                    "picture": jira_user.get("avatarUrls", {}).get("48x48", ""),
                }
                logger.info(f"User info from Jira myself: {u.get('email', 'unknown')}")
            except Exception as jira_err:
                logger.warning(f"Jira myself also failed ({jira_err}), using site name as identity")
                u = {
                    "name":    sites[0].get("name", "Jira User"),
                    "email":   "",
                    "picture": "",
                }

        session.permanent = True
        session["oauth_token"] = token
        session["user"] = {
            "name":   u.get("name", u.get("displayName", "")),
            "email":  u.get("email", ""),
            "avatar": u.get("picture", ""),
        }
        session["jira_sites"] = sites

        # If only one site, auto-select it
        if len(sites) == 1:
            session["jira_cloud_id"]  = sites[0]["id"]
            session["jira_site_name"] = sites[0]["name"]
            session["jira_site_url"]  = sites[0].get("url", "")
            ensure_selected_project()
            return redirect(url_for("index"))

        # Multiple sites — let user pick
        return redirect(url_for("select_site"))

    except Exception as e:
        logger.error(f"OAuth callback error: {e}")
        return render_template("login.html", error=f"Login failed: {str(e)}")


@app.route("/select-site")
@login_required
def select_site():
    """Let user pick which Jira site to use (if they have multiple)."""
    sites = session.get("jira_sites", [])
    return render_template("select_site.html", sites=sites, user=current_user())

@app.route("/dashboard")
def dashboard():
    return render_template("dashboard.html")


@app.route("/api/burndown")
def api_burndown():
    from datetime import datetime, timedelta
    # 10-day sample dataset (most recent last)
    days = 10
    dates = [(datetime.utcnow() - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d") for i in range(days)]
    total = 50
    ideal = [round(total * (1 - i / (days - 1))) for i in range(days)]
    # simple synthetic actuals (no Jira required)
    actual = [max(0, ideal[i] - (i % 3) * 2) for i in range(days)]
    return jsonify({"dates": dates, "ideal": ideal, "actual": actual})

@app.route("/select-site", methods=["POST"])
@login_required
def select_site_post():
    """Save selected Jira site."""
    cloud_id  = request.form.get("cloud_id")
    sites     = session.get("jira_sites", [])
    site      = next((s for s in sites if s["id"] == cloud_id), None)
    if not site:
        return redirect(url_for("select_site"))

    session["jira_cloud_id"]  = cloud_id
    session["jira_site_name"] = site["name"]
    session["jira_site_url"]  = site.get("url", "")
    ensure_selected_project()
    return redirect(url_for("index"))


@app.route("/select-project")
@login_required
def select_project():
    """Let user pick which Jira project to work with."""
    token    = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    projects = fetch_user_projects(cloud_id, token.get("access_token", ""))
    return render_template("select_project.html", projects=projects,
                           user=current_user(), site=session.get("jira_site_name"))


@app.route("/select-project", methods=["POST"])
@login_required
def select_project_post():
    """Save selected project."""
    project_key  = request.form.get("project_key", "").upper().strip()
    project_name = request.form.get("project_name", "")
    if not project_key:
        return redirect(url_for("select_project"))
    session["project_key"]  = project_key
    session["project_name"] = project_name
    return redirect(url_for("index"))


@app.route("/api/project/select", methods=["POST"])
@login_required
def select_project_api():
    payload = request.get_json() or {}
    requested_key = (payload.get("project_key") or "").upper().strip()
    if not requested_key:
        return jsonify({"error": "project_key required"}), 400

    token = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    projects = fetch_user_projects(cloud_id, token.get("access_token", ""))
    selected = next((p for p in projects if (p.get("key") or "").upper() == requested_key), None)
    if not selected:
        return jsonify({"error": f"Project {requested_key} not found or no access"}), 404

    session["project_key"] = selected.get("key", requested_key)
    session["project_name"] = selected.get("name", "")
    clear_user_cache()
    return jsonify({
        "message": "Project selected",
        "project_key": session.get("project_key", ""),
        "project_name": session.get("project_name", ""),
    })


@app.route("/auth/status")
def auth_status():
    if is_authenticated():
        cfg = get_jira_auth()
        return jsonify({
            "connected":    True,
            "auth_type":    cfg["auth_type"] if cfg else "unknown",
            "user":         current_user(),
            "site_name":    session.get("jira_site_name", ""),
            "project_key":  get_user_project_key(),
            "project_name": session.get("project_name", ""),
        })
    return jsonify({"connected": False})


@app.route("/auth/logout")
def auth_logout():
    # Clear per-user dashboards from memory on logout
    uid = user_id()
    if uid and uid in _user_dashboards:
        del _user_dashboards[uid]
        logger.info(f"Dashboards cleared for user {uid}")
    # Clear user cache
    clear_user_cache()
    session.clear()
    response = redirect(url_for("login_page"))
    response.delete_cookie("oauth_state")
    return response


# ══════════════════════════════════════════════════════════════════════════════
# App Routes (all protected)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def index():
    if session.get("jira_cloud_id") and not get_user_project_key():
        ensure_selected_project()
    return render_template("index.html",
        user=current_user(),
        site=session.get("jira_site_name", ""),
        project_key=get_user_project_key(),
        project_name=session.get("project_name", ""),
    )


@app.route("/api/generate", methods=["POST"])
@login_required
def generate_dashboard():
    data   = request.get_json()
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return jsonify({"error": "Prompt is required"}), 400

    result = process_prompt(prompt)

    if result.get("error") and not result.get("needs_project"):
        return jsonify(result), 400

    if result.get("dashboard_config"):
        dash_id = result["dashboard_config"]["dashboard_id"]
        save_user_dashboard(dash_id, {
            "id": dash_id,
            "prompt": prompt,
            "dashboard_config": result["dashboard_config"],
            "created_at": datetime.now().isoformat(),
        })
        result["dashboard_id"] = dash_id

    elif result.get("chart_config"):
        dash_id = result["chart_config"]["chart_id"]
        save_user_dashboard(dash_id, {
            "id":           dash_id,
            "prompt":       prompt,
            "chart_config": result["chart_config"],
            "jira_task":    result.get("jira_task"),
            "created_at":   datetime.now().isoformat(),
        })
        result["dashboard_id"] = dash_id

    return jsonify(result)


@app.route("/api/dashboards", methods=["GET"])
@login_required
def list_dashboards():
    dashes = get_user_dashboards()
    return jsonify({"dashboards": list(dashes.values()), "count": len(dashes)})


@app.route("/api/dashboard/<dashboard_id>", methods=["GET"])
@login_required
def get_dashboard(dashboard_id):
    dashes = get_user_dashboards()
    if dashboard_id not in dashes:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dashes[dashboard_id])


@app.route("/api/dashboard/<dashboard_id>", methods=["PUT"])
@login_required
def update_dashboard(dashboard_id):
    dashes = get_user_dashboards()
    if dashboard_id not in dashes:
        return jsonify({"error": "Not found"}), 404
    prompt = (request.get_json() or {}).get("prompt", "").strip()
    if not prompt:
        return jsonify({"error": "Prompt required"}), 400
    result = process_prompt(prompt)
    updates = {
        "prompt": prompt,
        "updated_at": datetime.now().isoformat(),
    }
    if result.get("dashboard_config"):
        updates["dashboard_config"] = result.get("dashboard_config")
        updates["chart_config"] = None
    else:
        updates["chart_config"] = result.get("chart_config")
        updates["dashboard_config"] = None
    dashes[dashboard_id].update(updates)
    save_user_dashboard(dashboard_id, dashes[dashboard_id])
    return jsonify({"message": "Updated", "dashboard": dashes[dashboard_id]})


@app.route("/api/jira/tasks", methods=["GET"])
@login_required
def get_jira_tasks():
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503
    project_key = get_user_project_key()
    try:
        data   = jira_get_cached("search/jql", params={
            "jql":        f"project={project_key} ORDER BY created DESC",
            "maxResults": 50,
            "fields":     "summary,status,assignee,priority,created",
        })
        cfg    = get_jira_auth()
        site   = session.get("jira_site_name", "")
        server = f"https://{site}.atlassian.net" if site else os.getenv("JIRA_SERVER", "")
        tasks  = []
        for issue in data.get("issues", []):
            f = issue["fields"]
            tasks.append({
                "task_id":  issue["key"],
                "summary":  f.get("summary", ""),
                "status":   f["status"]["name"],
                "assignee": f["assignee"]["displayName"] if f.get("assignee") else "Unassigned",
                "priority": f["priority"]["name"] if f.get("priority") else "None",
                "created":  f.get("created", ""),
                "url":      f"{server}/browse/{issue['key']}",
            })
        return jsonify({"tasks": tasks, "count": len(tasks)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/projects", methods=["GET"])
@login_required
def get_user_projects_api():
    """Return list of projects for the current user's Jira site."""
    token    = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    if not cloud_id:
        return jsonify({"error": "No Jira site selected"}), 400
    projects = fetch_user_projects(cloud_id, token.get("access_token", ""))
    return jsonify({"projects": projects})


@app.route("/api/jira/releases", methods=["GET"])
@login_required
def get_jira_releases():
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503

    requested_project_key = (request.args.get("project_key") or "").strip().upper()
    project_key = requested_project_key or get_user_project_key()
    if not project_key:
        return jsonify({"error": "No project selected"}), 400

    try:
        versions = []

        # Try the paginated project versions API first.
        try:
            paged_versions = jira_get_cached(f"project/{project_key}/version")
            if isinstance(paged_versions, dict):
                versions = paged_versions.get("values", []) or []
            elif isinstance(paged_versions, list):
                versions = paged_versions
        except Exception as paged_err:
            logger.warning(f"Paged release fetch failed for {project_key}: {paged_err}")

        # Fallback to the legacy versions endpoint when needed.
        if not versions:
            try:
                legacy_versions = jira_get_cached(f"project/{project_key}/versions")
                if isinstance(legacy_versions, list):
                    versions = legacy_versions
                elif isinstance(legacy_versions, dict):
                    versions = legacy_versions.get("values", []) or []
            except Exception as legacy_err:
                logger.warning(f"Legacy release fetch failed for {project_key}: {legacy_err}")

        # Final fallback: infer releases from issue fixVersions when project version APIs are empty.
        if not versions:
            try:
                release_issue_data = jira_get_cached("search/jql", params={
                    "jql": f"project={project_key} AND fixVersion IS NOT EMPTY ORDER BY updated DESC",
                    "maxResults": 200,
                    "fields": "fixVersions,status",
                })
                version_map: Dict[str, Dict[str, Any]] = {}
                for issue in release_issue_data.get("issues", []):
                    fields = issue.get("fields", {})
                    is_done = is_done_status((fields.get("status") or {}).get("name", ""))
                    for version in fields.get("fixVersions", []) or []:
                        version_id = version.get("id") or version.get("name") or str(uuid.uuid4())
                        record = version_map.setdefault(version_id, {
                            "id": version.get("id"),
                            "name": version.get("name") or "Unnamed release",
                            "released": bool(version.get("released")),
                            "archived": bool(version.get("archived")),
                            "releaseDate": version.get("releaseDate") or "",
                            "startDate": version.get("startDate") or "",
                            "description": version.get("description") or "",
                            "issuesCount": 0,
                            "issuesFixedCount": 0,
                        })
                        record["issuesCount"] += 1
                        if is_done:
                            record["issuesFixedCount"] += 1
                versions = list(version_map.values())
            except Exception as fix_version_err:
                logger.warning(f"fixVersions fallback failed for {project_key}: {fix_version_err}")

        releases = []
        for version in versions:
            version_id = version.get("id")
            related_counts = {}
            if version_id:
                try:
                    related_counts = jira_get_cached(f"version/{version_id}/relatedIssueCounts")
                except Exception as related_err:
                    logger.warning(f"Release related issue counts unavailable for version {version_id}: {related_err}")

            issues_fixed = int(related_counts.get("issuesFixedCount") or version.get("issuesFixedCount") or 0)
            issues_total = int(related_counts.get("issuesCount") or version.get("issuesCount") or issues_fixed or 0)
            progress_percent = 0
            if issues_total > 0:
                progress_percent = round((issues_fixed / issues_total) * 100)

            release_date = version.get("releaseDate") or ""
            start_date = version.get("startDate") or ""
            released = bool(version.get("released"))
            archived = bool(version.get("archived"))
            status = "Released" if released else ("Archived" if archived else "Unreleased")

            releases.append({
                "id": version_id,
                "name": version.get("name") or "Unnamed release",
                "status": status,
                "released": released,
                "archived": archived,
                "start_date": start_date,
                "release_date": release_date,
                "description": version.get("description") or "",
                "issues_total": issues_total,
                "issues_fixed": issues_fixed,
                "progress_percent": progress_percent,
            })

        def sort_key(item):
            release_date = item.get("release_date") or "9999-12-31"
            return (item.get("released", False), release_date, item.get("name", ""))

        releases.sort(key=sort_key)
        return jsonify({"releases": releases[:8], "count": len(releases)})
    except Exception as e:
        logger.error(f"Release fetch failed for {project_key}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/project/intelligence", methods=["GET"])
@login_required
def get_project_intelligence():
    """Return predictive project health and recommended actions."""
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503

    project_key = get_user_project_key()
    if not project_key:
        return jsonify({"error": "No project selected"}), 400

    try:
        intelligence = build_project_intelligence(project_key)
        return jsonify(intelligence)
    except Exception as e:
        logger.error(f"Project intelligence failed for {project_key}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/tasks", methods=["POST"])
@login_required
def create_jira_task():
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503
    project_key = get_user_project_key()
    data        = request.get_json()
    try:
        payload = {"fields": {
            "project":     {"key": project_key},
            "summary":     data.get("summary", "New Task"),
            "issuetype":   {"name": "Task"},
            "priority":    {"name": data.get("priority", "Medium")},
            "description": {"type": "doc", "version": 1, "content": [
                {"type": "paragraph", "content": [{"type": "text", "text": data.get("description", "")}]}
            ]},
        }}
        result    = jira_post("issue", payload)
        clear_user_cache()
        issue_key = result.get("key")
        site      = session.get("jira_site_name", "")
        server    = f"https://{site}.atlassian.net" if site else os.getenv("JIRA_SERVER", "")
        return jsonify({"message": "Task created", "task": {"task_id": issue_key, "url": f"{server}/browse/{issue_key}"}}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/tasks/<task_id>/edit-data", methods=["GET"])
@login_required
def get_jira_task_edit_data(task_id):
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503

    try:
        issue = jira_get(f"issue/{task_id}", params={"fields": "summary,status,assignee,priority"})
        fields = issue.get("fields", {})

        current = {
            "summary": fields.get("summary", ""),
            "status": fields.get("status", {}).get("name", "") if fields.get("status") else "",
            "assignee": {
                "accountId": fields.get("assignee", {}).get("accountId", "") if fields.get("assignee") else "",
                "displayName": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
            },
            "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
        }

        transitions_data = jira_get(f"issue/{task_id}/transitions")
        transitions = [
            {"id": t.get("id", ""), "name": t.get("name", "")}
            for t in transitions_data.get("transitions", [])
        ]

        assignable_raw = jira_get("user/assignable/search", params={"issueKey": task_id, "maxResults": 100})
        assignable_users = assignable_raw if isinstance(assignable_raw, list) else []
        assignees = [
            {
                "accountId": u.get("accountId", ""),
                "displayName": u.get("displayName", "Unknown User"),
            }
            for u in assignable_users
            if u.get("accountId")
        ]

        priority_names = []
        try:
            editmeta = jira_get(f"issue/{task_id}/editmeta")
            allowed_values = (
                editmeta.get("fields", {})
                .get("priority", {})
                .get("allowedValues", [])
            )
            priority_names = [p.get("name", "") for p in allowed_values if p.get("name")]
        except Exception:
            # Keep endpoint resilient even if editmeta is restricted.
            priority_names = []

        if not priority_names:
            priority_names = ["Lowest", "Low", "Medium", "High", "Highest"]

        if current["priority"] and current["priority"] not in priority_names:
            priority_names.insert(0, current["priority"])

        return jsonify({
            "task_id": task_id,
            "current": current,
            "assignees": assignees,
            "priorities": priority_names,
            "statuses": transitions,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/tasks/<task_id>", methods=["PUT"])
@login_required
def update_jira_task(task_id):
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503
    data = request.get_json()
    try:
        fields = {}
        if "summary" in data:
            fields["summary"] = data.get("summary", "")

        if "assignee_account_id" in data:
            account_id = (data.get("assignee_account_id") or "").strip()
            fields["assignee"] = {"accountId": account_id} if account_id else None

        if "priority" in data:    fields["priority"]    = {"name": data["priority"]}
        if "description" in data: fields["description"] = {"type": "doc", "version": 1, "content": [
            {"type": "paragraph", "content": [{"type": "text", "text": data["description"]}]}]}

        if fields:
            jira_put(f"issue/{task_id}", {"fields": fields})

        if "status_id" in data and data.get("status_id"):
            jira_post(f"issue/{task_id}/transitions", {"transition": {"id": str(data.get("status_id"))}})
        if "status" in data:
            trans   = jira_get(f"issue/{task_id}/transitions")
            matched = next((t for t in trans.get("transitions", []) if t["name"].lower() == data["status"].lower()), None)
            if matched:
                jira_post(f"issue/{task_id}/transitions", {"transition": {"id": matched["id"]}})
        clear_user_cache()
        return jsonify({"message": f"{task_id} updated"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/<dashboard_id>", methods=["GET"])
@login_required
def export_dashboard(dashboard_id):
    dashes = get_user_dashboards()
    if dashboard_id not in dashes:
        return jsonify({"error": "Not found"}), 404
    
    dash = dashes[dashboard_id]
    cfg = dash.get("chart_config", {})
    project_key = get_user_project_key()

    # Re-fetch real Jira data for export when possible to avoid exporting sample fallback.
    prompt = (dash.get("prompt") or "").strip()
    export_cfg = dict(cfg)
    fresh_chart_data = ensure_jira_chart_data_for_export(prompt, project_key, {})

    if fresh_chart_data:
        chart_type = export_cfg.get("chart_type") or detect_chart_type(prompt or "")
        title = fresh_chart_data.get("title") or export_cfg.get("title") or "Jira Chart"
        labels = fresh_chart_data.get("labels", [])
        values = fresh_chart_data.get("values", [])

        export_cfg["title"] = title
        export_cfg["labels"] = labels
        export_cfg["values"] = values
        export_cfg["image_base64"] = generate_chart_b64(chart_type, title, labels, values)
        export_cfg["data_source"] = "jira_mcp"

        chart_data = {
            "title": title,
            "labels": labels,
            "values": values,
        }
    else:
        # If no fresh Jira data is available and current chart is known sample fallback,
        # fail explicitly instead of exporting misleading data.
        if (export_cfg.get("data_source") == "sample_fallback") or (export_cfg.get("title") == "Sample Distribution"):
            return jsonify({"error": "No real Jira data available for export. Please refresh Jira connection or regenerate dashboard."}), 400

        chart_data = {
            "title": export_cfg.get("title", "Chart"),
            "labels": export_cfg.get("labels", []),
            "values": export_cfg.get("values", []),
        }
    
    # Create Excel with data table and chart image
    excel_buf = create_excel_with_chart(chart_data, export_cfg, project_key)
    
    return send_file(excel_buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"dashboard_{dashboard_id}.xlsx")


@app.route("/api/metrics", methods=["GET"])
@login_required
def get_metrics():
    dashes = get_user_dashboards()
    types  = ["pie", "bar", "line", "metrics"]
    return jsonify({
        "total_dashboards": len(dashes),
        "chart_types": {t: sum(1 for d in dashes.values() if d.get("chart_config", {}).get("chart_type") == t) for t in types},
    })


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("DEBUG", "true").lower() in {"1", "true", "yes", "on"}
    print("=" * 60)
    print("  Dashboard Generator — Multi-Tenant Edition")
    print("=" * 60)
    print(f"  OAuth Client : {os.getenv('JIRA_CLIENT_ID','Not set')}")
    print(f"  LLM Model    : {os.getenv('QWEN_MODEL','Not set')}")
    print(f"  Login URL    : http://localhost:{port}/login")
    print("=" * 60)
    app.run(debug=debug, host="0.0.0.0", port=port)
