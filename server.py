"""
Dynamic Dashboard Generator — Multi-Tenant Edition
- Every user logs in with their OWN Atlassian account
- Supports any Jira instance (balajiselliappan.atlassian.net, ABCCorp.atlassian.net, etc.)
- All data is isolated per user session
- No shared state between users
"""

import os
import re
import json
import base64
import uuid
import logging
import time
import hashlib
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Any, Optional
from functools import wraps

from flask import (
    Flask, render_template, request, jsonify,
    send_file, redirect, session, url_for, abort
)
from flask_cors import CORS
from dotenv import load_dotenv
from authlib.integrations.flask_client import OAuth

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import requests
from requests.auth import HTTPBasicAuth

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage

from mcp_client import call_jira_tool, list_jira_tools, MCPClientError
from flask_session import Session

# ── Env + logging ─────────────────────────────────────────────────────────────
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

JIRA_OAUTH_SCOPES = "read:jira-work write:jira-work read:jira-user read:board-scope:jira-software read:sprint:jira-software offline_access"
#JIRA_OAUTH_SCOPES = "read:jira-work write:jira-work read:jira-user read:board-scope:jira-software read:sprint:jira-software offline_access"JIRA_OAUTH_SCOPES = "read:jira-work write:jira-work read:jira-user offline_access"

# ── Flask ─────────────────────────────────────────────────────────────────────
app = Flask(__name__)
# IMPORTANT: Must be a fixed key — os.urandom() changes on restart and breaks OAuth state
_default_secret = "jira-dashboard-secret-key-change-this-in-production-2026"
app.secret_key = os.getenv("FLASK_SECRET_KEY", _default_secret)
app.config["PERMANENT_SESSION_LIFETIME"] = 28800  # 8 hours
app.config["SESSION_TYPE"]            = "filesystem"
app.config["SESSION_FILE_DIR"]        = "/tmp/jira_dashboard_sessions"
app.config["SESSION_FILE_THRESHOLD"]  = 100
CORS(app)

# Session config — critical for OAuth state to survive the redirect round-trip
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"    # allow cross-site redirect callbacks
app.config["SESSION_COOKIE_SECURE"]   = False     # False for localhost (True in production)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_NAME"]     = "jira_dashboard_session" 

# ── OAuth ─────────────────────────────────────────────────────────────────────
# Init server-side session (stores OAuth state server-side, not in cookie)
import os as _os
_os.makedirs("/tmp/jira_dashboard_sessions", exist_ok=True)
Session(app)

oauth = OAuth(app)
atlassian = oauth.register(
    name="atlassian",
    client_id=os.getenv("JIRA_CLIENT_ID"),
    client_secret=os.getenv("JIRA_CLIENT_SECRET"),
    authorize_url="https://auth.atlassian.com/authorize",
    access_token_url="https://auth.atlassian.com/oauth/token",
    client_kwargs={
        "scope": JIRA_OAUTH_SCOPES,
        "audience": "api.atlassian.com",
        "prompt": "consent",
    },
)

# ── Per-user dashboard store: {user_id: {dash_id: dash_data}} ────────────────
_user_dashboards: Dict[str, Dict] = {}

# ── Per-user Jira cache: {cache_key: (data, timestamp)} ─────────────────────
_user_cache: Dict[str, Any] = {}
CACHE_TTL = 60


# ══════════════════════════════════════════════════════════════════════════════
# Session helpers — everything scoped to current user
# ══════════════════════════════════════════════════════════════════════════════

def current_user() -> Optional[Dict]:
    """Return current logged-in user info from session."""
    return session.get("user")


def user_id() -> Optional[str]:
    """Return a stable user ID (hashed email)."""
    user = current_user()
    if not user:
        return None
    return hashlib.md5(user.get("email", "").encode()).hexdigest()


def is_authenticated() -> bool:
    """Check if user is logged in and token exists."""
    return bool(session.get("oauth_token") and session.get("jira_cloud_id") and current_user())


def get_user_project_key() -> str:
    """Get the Jira project key selected by the current user."""
    return session.get("project_key", "")


def get_user_dashboards() -> Dict:
    """Get dashboards for the current user only."""
    uid = user_id()
    if not uid:
        return {}
    return _user_dashboards.get(uid, {})


def save_user_dashboard(dash_id: str, data: Dict):
    """Save a dashboard for the current user."""
    uid = user_id()
    if not uid:
        return
    if uid not in _user_dashboards:
        _user_dashboards[uid] = {}
    _user_dashboards[uid][dash_id] = data


# ── login_required decorator ──────────────────────────────────────────────────
def login_required(f):
    """Redirect to login if user not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not is_authenticated():
            if request.is_json:
                return jsonify({"error": "Not authenticated", "redirect": "/login"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
# Jira Auth — uses current user's OAuth token
# ══════════════════════════════════════════════════════════════════════════════

def get_jira_auth() -> Optional[Dict]:
    """
    Return Jira connection config for the CURRENT USER.
    Uses their OAuth token from session.
    Falls back to .env API key only if no OAuth token present.
    """
    token    = session.get("oauth_token")
    cloud_id = session.get("jira_cloud_id")

    if token and cloud_id:
        return {
            "base_url":  f"https://api.atlassian.com/ex/jira/{cloud_id}",
            "auth":      None,
            "headers":   {
                "Authorization": f"Bearer {token.get('access_token')}",
                "Accept":        "application/json",
                "Content-Type":  "application/json",
            },
            "auth_type": "oauth",
            "site":      session.get("jira_site_name", ""),
        }

    # API key fallback
    server    = os.getenv("JIRA_SERVER", "").rstrip("/")
    email     = os.getenv("JIRA_USER_EMAIL")
    api_token = os.getenv("JIRA_API_TOKEN")
    if all([server, email, api_token]):
        return {
            "base_url":  server,
            "auth":      HTTPBasicAuth(email, api_token),
            "headers":   {"Accept": "application/json", "Content-Type": "application/json"},
            "auth_type": "apikey",
            "site":      server.replace("https://", ""),
        }

    return None


def get_mcp_auth_context() -> Optional[Dict]:
    """Return session-scoped auth context passed to MCP tool calls."""
    token = session.get("oauth_token")
    cloud_id = session.get("jira_cloud_id")

    if token and cloud_id:
        return {
            "auth_type": "oauth",
            "cloud_id": cloud_id,
            "access_token": token.get("access_token", ""),
            "site": session.get("jira_site_name", ""),
        }

    server = os.getenv("JIRA_SERVER", "").rstrip("/")
    email = os.getenv("JIRA_USER_EMAIL")
    api_token = os.getenv("JIRA_API_TOKEN")
    if all([server, email, api_token]):
        return {
            "auth_type": "apikey",
            "server": server,
            "email": email,
            "api_token": api_token,
        }

    return None


def jira_get(path: str, params: dict = None) -> dict:
    cfg  = get_jira_auth()
    if not cfg:
        raise ValueError("Not authenticated with Jira")
    logger.info(f"[{cfg['auth_type'].upper()}] GET {path}")
    url  = f"{cfg['base_url']}/rest/api/3/{path.lstrip('/')}"
    resp = requests.get(url, auth=cfg.get("auth"), headers=cfg["headers"], params=params)
    resp.raise_for_status()
    return resp.json()


def jira_get_cached(path: str, params: dict = None) -> dict:
    """Cached version of jira_get — 60s TTL, scoped per user."""
    uid = user_id() or "anon"
    key = f"{uid}:{path}:{json.dumps(params or {}, sort_keys=True)}"
    now = time.time()
    if key in _user_cache:
        data, ts = _user_cache[key]
        if now - ts < CACHE_TTL:
            logger.info(f"[CACHE HIT] {path}")
            return data
    data = jira_get(path, params)
    _user_cache[key] = (data, now)
    return data


def jira_post(path: str, payload: dict) -> dict:
    cfg  = get_jira_auth()
    if not cfg:
        raise ValueError("Not authenticated with Jira")
    logger.info(f"[{cfg['auth_type'].upper()}] POST {path}")
    url  = f"{cfg['base_url']}/rest/api/3/{path.lstrip('/')}"
    resp = requests.post(url, auth=cfg.get("auth"), headers=cfg["headers"], json=payload)
    resp.raise_for_status()
    return resp.json() if resp.content else {}


def jira_put(path: str, payload: dict) -> dict:
    cfg  = get_jira_auth()
    if not cfg:
        raise ValueError("Not authenticated with Jira")
    url  = f"{cfg['base_url']}/rest/api/3/{path.lstrip('/')}"
    resp = requests.put(url, auth=cfg.get("auth"), headers=cfg["headers"], json=payload)
    resp.raise_for_status()
    return resp.json() if resp.content else {}


def clear_user_cache():
    """Clear cache entries for the current user."""
    uid = user_id() or "anon"
    keys_to_delete = [k for k in _user_cache if k.startswith(f"{uid}:")]
    for k in keys_to_delete:
        del _user_cache[k]
    logger.info(f"Cache cleared for user {uid}")


# ══════════════════════════════════════════════════════════════════════════════
# Jira site helpers
# ══════════════════════════════════════════════════════════════════════════════

def fetch_accessible_sites(access_token: str) -> List[Dict]:
    """Fetch all Jira sites the user has access to."""
    try:
        resp = requests.get(
            "https://api.atlassian.com/oauth/token/accessible-resources",
            headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.error(f"fetch_accessible_sites failed: {e}")
        return []


def fetch_user_projects(cloud_id: str, access_token: str) -> List[Dict]:
    """Fetch all projects from the user's Jira site."""
    try:
        url  = f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/api/3/project"
        resp = requests.get(
            url,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Accept":        "application/json",
            }
        )
        resp.raise_for_status()
        projects = resp.json()
        return [{"key": p["key"], "name": p["name"]} for p in projects]
    except Exception as e:
        logger.error(f"fetch_user_projects failed: {e}")
        return []


def ensure_selected_project() -> bool:
    """Ensure a valid project is selected in session for the current Jira site."""
    token = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    access_token = token.get("access_token", "")
    if not cloud_id or not access_token:
        return False

    projects = fetch_user_projects(cloud_id, access_token)
    if not projects:
        return False

    current_key = (session.get("project_key") or "").upper().strip()
    current = next((p for p in projects if (p.get("key") or "").upper() == current_key), None)

    if current:
        session["project_key"] = current.get("key", "")
        session["project_name"] = current.get("name", "")
        return True

    first = projects[0]
    session["project_key"] = first.get("key", "")
    session["project_name"] = first.get("name", "")
    return True


# ══════════════════════════════════════════════════════════════════════════════
# LLM setup
# ══════════════════════════════════════════════════════════════════════════════

def get_llm(fast: bool = False) -> Optional[ChatOpenAI]:
    api_key  = os.getenv("QWEN_API_KEY", "")
    base_url = os.getenv("QWEN_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
    model    = os.getenv("QWEN_MODEL_FAST" if fast else "QWEN_MODEL", "qwen-plus")

    if not api_key:
        return None
    try:
        return ChatOpenAI(
            model=model, temperature=0.1,
            api_key=api_key, base_url=base_url,
            request_timeout=30, max_retries=1,
        )
    except Exception as e:
        logger.error(f"LLM init failed: {e}")
        return None


llm      = get_llm(fast=False)
llm_fast = get_llm(fast=True)


# ══════════════════════════════════════════════════════════════════════════════
# Fast dispatch — skip LLM for obvious queries
# ══════════════════════════════════════════════════════════════════════════════

def fast_dispatch(prompt_lower: str, project_key: str) -> Optional[Dict]:
    auth_context = get_mcp_auth_context()
    if not auth_context:
        return None

    def _mcp_call(name: str, args: Dict[str, Any]) -> Dict:
        try:
            merged = {"auth_context": auth_context, **args}
            return call_jira_tool(name, merged)
        except MCPClientError as e:
            logger.warning(f"FAST MCP call failed ({name}): {e}")
            return {"error": str(e)}

    if any(p in prompt_lower for p in ["list sprint", "show sprint", "get sprint", "all sprint", "sprints"]):
        logger.info("[FAST] jira_get_sprints")
        result = _mcp_call("jira_get_sprints", {"project_key": project_key})
        if result.get("error"):
            return {"error": result["error"], "response": result["error"]}
        sprints = result.get("sprints", [])
        return {"sprints": sprints, "result_type": "sprints",
                "response": f"Found {len(sprints)} sprint(s)", "jira_results": sprints}

    if any(p in prompt_lower for p in ["my issue", "my task", "assigned to me", "what am i working"]):
        logger.info("[FAST] jira_get_my_issues")
        result = _mcp_call("jira_get_my_issues", {"project_key": project_key})
        issues = result.get("issues", [])
        return {"issues": issues, "result_type": "issues",
                "response": f"Found {len(issues)} issue(s) assigned to you", "jira_results": issues}

    issue_match = re.search(r"[A-Z]+-[0-9]+", prompt_lower.upper())
    if issue_match and any(p in prompt_lower for p in ["tell me", "details", "about", "show", "status of"]):
        logger.info(f"[FAST] jira_get_issue({issue_match.group()})")
        result = _mcp_call("jira_get_issue", {"issue_key": issue_match.group()})
        return {"result_type": "issue_detail", "jira_results": [result],
                "response": f"Details for {issue_match.group()}"}

    if any(p in prompt_lower for p in ["list issue", "show issue", "all issue", "show all", "list all"]):
        # Only treat "in <project>" as a project hint when it appears at the end,
        # so phrases like "in progress" do not get parsed as project keys.
        project_hint_match = re.search(r"\bin\s+([a-z][a-z0-9]+)\s*$", prompt_lower)
        effective_project_key = project_key
        if project_hint_match:
            hinted = project_hint_match.group(1).upper().strip()
            if 1 <= len(hinted) <= 12:
                effective_project_key = hinted

        sprint_match = re.search(r"sprint\s*(\d+)", prompt_lower)
        jql = f"project={effective_project_key}"
        if sprint_match:
            jql += f' AND sprint = "{effective_project_key} Sprint {sprint_match.group(1)}"'
        if "in progress" in prompt_lower:   jql += ' AND status = "In Progress"'
        elif "to do"      in prompt_lower:  jql += ' AND status = "To Do"'
        elif "done"       in prompt_lower:  jql += ' AND status = "Done"'
        jql += " ORDER BY created DESC"
        logger.info(f"[FAST] direct jira_search_issues({jql})")
        try:
            data = jira_get_cached("search/jql", params={
                "jql": jql,
                "maxResults": 50,
                "fields": "summary,status,assignee,priority,created,updated,issuetype",
            })
            raw_issues = data.get("issues", [])
            issues = []
            for item in raw_issues:
                fields = item.get("fields", {})
                issues.append({
                    "key": item.get("key", ""),
                    "summary": fields.get("summary", ""),
                    "status": fields.get("status", {}).get("name", ""),
                    "assignee": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
                    "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
                    "type": fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else "",
                    "created": fields.get("created", "")[:10] if fields.get("created") else "",
                    "updated": fields.get("updated", "")[:10] if fields.get("updated") else "",
                })

            total = data.get("total", len(issues))
            return {
                "issues": issues,
                "result_type": "issues",
                "response": f"Found {total} issue(s)",
                "jira_results": issues,
            }
        except Exception as direct_err:
            logger.warning(f"[FAST] direct jira_search_issues failed, trying MCP fallback: {direct_err}")
            result = _mcp_call("jira_search_issues", {"jql": jql, "max_results": 50})
            if result.get("error"):
                return {
                    "error": f"Issue lookup failed: {direct_err} | MCP fallback failed: {result.get('error')}",
                    "response": "Could not fetch issues. Please retry.",
                }
            issues = result.get("issues", [])
            return {
                "issues": issues,
                "result_type": "issues",
                "response": f"Found {result.get('total', len(issues))} issue(s)",
                "jira_results": issues,
            }

    if any(p in prompt_lower for p in ["who is working", "who has", "assignee breakdown"]):
        logger.info("[FAST] jira_get_project_summary (assignees)")
        result = _mcp_call("jira_get_project_summary", {"project_key": project_key})
        counts  = result.get("assignee_counts", {})
        results = [{"name": k, "count": v} for k, v in counts.items()]
        return {"result_type": "assignees", "jira_results": results,
                "response": f"Found {len(results)} assignees"}

    return None


# ══════════════════════════════════════════════════════════════════════════════
# MCP Agent
# ══════════════════════════════════════════════════════════════════════════════

def run_jira_agent(prompt: str, project_key: str) -> Dict:
    if not llm:
        return {"error": "LLM not configured", "raw_response": ""}

    auth_context = get_mcp_auth_context()
    if not auth_context:
        return {"error": "Jira auth context is missing", "raw_response": ""}

    try:
        tool_specs = list_jira_tools()
    except Exception as e:
        logger.error(f"MCP tools/list failed: {e}")
        return {"error": f"MCP unavailable: {e}", "raw_response": ""}

    tool_map = {t.get("name"): t for t in tool_specs if t.get("name")}
    tool_descriptions = "\n".join([
        f"- {t.get('name')}: {(t.get('description') or '').strip().split(chr(10))[0]}"
        for t in tool_specs
    ])

    system_prompt = f"""You are a Jira assistant. Current project key: {project_key}

Available tools:
{tool_descriptions}

To use a tool respond ONLY with JSON:
{{"tool": "tool_name", "args": {{"arg1": "value1"}}}}

To give final answer:
{{"result": {{...}}, "summary": "explanation"}}

Always use project key: {project_key}
"""
    messages = [SystemMessage(content=system_prompt), HumanMessage(content=prompt)]
    tool_results = []

    for i in range(5):
        try:
            response = llm.invoke(messages)
            raw_text = response.content.strip()
            logger.info(f"Agent iter {i+1}: {raw_text[:200]}")

            json_match = re.search(r"\{.*\}", raw_text, re.DOTALL)
            if not json_match:
                return {"raw_response": raw_text}

            parsed = json.loads(json_match.group())

            if "tool" in parsed:
                tool_name = parsed.get("tool")
                tool_args = parsed.get("args", {})
                if tool_name not in tool_map:
                    return {"raw_response": raw_text}
                tool_args = tool_args or {}
                tool_args["auth_context"] = auth_context
                if tool_name in {"jira_get_sprints", "jira_get_project_summary", "jira_get_my_issues", "jira_create_issue"}:
                    tool_args.setdefault("project_key", project_key)
                logger.info(f"MCP tool call: {tool_name}({tool_args})")
                try:
                    tool_result = call_jira_tool(tool_name, tool_args)
                    tool_results.append({"tool": tool_name, "result": tool_result})
                    messages.append(response)
                    messages.append(HumanMessage(
                        content=f"Tool {tool_name} returned: {json.dumps(tool_result)}\n\nNow give the final result as JSON."
                    ))
                except Exception as e:
                    return {"raw_response": f"Tool {tool_name} failed: {e}"}

            elif "result" in parsed:
                result = parsed["result"]
                result["raw_response"] = parsed.get("summary", raw_text)
                result["tool_calls"]   = [t["tool"] for t in tool_results]
                return result
            else:
                return {"raw_response": raw_text}

        except Exception as e:
            logger.error(f"Agent iter {i+1} error: {e}")
            return {"error": str(e), "raw_response": ""}

    if tool_results:
        last = tool_results[-1]["result"]
        last["raw_response"] = f"Completed via {[t['tool'] for t in tool_results]}"
        return last
    return {"raw_response": "Could not complete the request"}


# ══════════════════════════════════════════════════════════════════════════════
# Chart generation
# ══════════════════════════════════════════════════════════════════════════════

def generate_chart_b64(chart_type: str, title: str, labels: list, values: list) -> str:
    palette = ["#0f9d8a", "#2f80ed", "#ff6b4a", "#f0b429", "#5b6cff", "#2eaf74"]
    labels = labels or ["No Data"]
    raw_values = values or [0]
    numeric_values = []
    for v in raw_values:
        try:
            numeric_values.append(float(v))
        except Exception:
            numeric_values.append(0.0)

    fig, ax = plt.subplots(figsize=(10, 6), dpi=150)
    fig.patch.set_facecolor("#f8fafc")
    ax.set_facecolor("#ffffff")

    colors = [palette[i % len(palette)] for i in range(max(len(labels), 1))]
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#d9e2ef")
    ax.spines["bottom"].set_color("#d9e2ef")

    if chart_type == "pie":
        total = sum(numeric_values)
        pie_values = numeric_values if total > 0 else [1]
        pie_labels = labels if total > 0 else ["No Data"]
        ax.pie(
            pie_values,
            labels=pie_labels,
            autopct="%1.1f%%" if total > 0 else None,
            colors=colors[:len(pie_labels)],
            startangle=90,
            wedgeprops={"width": 0.55, "edgecolor": "white", "linewidth": 2},
            textprops={"fontsize": 10, "color": "#24344d"},
        )
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=16)

    elif chart_type == "line":
        ax.plot(labels, numeric_values, marker="o", linewidth=3, markersize=7, color=palette[1])
        ax.fill_between(range(len(labels)), numeric_values, color=palette[1], alpha=0.14)
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=12)
        ax.set_xlabel("Categories", color="#66758f", fontsize=11)
        ax.set_ylabel("Values", color="#66758f", fontsize=11)
        ax.tick_params(axis="x", rotation=28, labelsize=10)
        ax.tick_params(axis="y", labelsize=10)
        ax.grid(axis="y", alpha=0.25, linestyle="--")

    elif chart_type == "metrics":
        ax.axis("off")
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=10)
        for i, (label, value) in enumerate(zip(labels, raw_values)):
            ax.text(
                0.5,
                0.84 - i * 0.18,
                f"{label}: {value}",
                ha="center",
                va="center",
                fontsize=15,
                fontweight="bold",
                color="#1f2f46",
                bbox={
                    "boxstyle": "round,pad=0.45",
                    "facecolor": colors[i % len(colors)],
                    "edgecolor": "none",
                    "alpha": 0.18,
                },
                transform=ax.transAxes,
            )

    else:
        bars = ax.bar(labels, numeric_values, color=colors[:len(labels)], edgecolor="none")
        ax.set_title(title, fontsize=16, fontweight="bold", color="#152238", pad=12)
        ax.set_xlabel("Categories", color="#66758f", fontsize=11)
        ax.set_ylabel("Values", color="#66758f", fontsize=11)
        ax.tick_params(axis="x", rotation=28, labelsize=10)
        ax.tick_params(axis="y", labelsize=10)
        ax.grid(axis="y", alpha=0.25, linestyle="--")
        y_max = max(numeric_values) if numeric_values else 0
        pad = max(y_max * 0.02, 0.5)
        for bar, value in zip(bars, raw_values):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + pad,
                str(value),
                ha="center",
                va="bottom",
                fontsize=10,
                color="#24344d",
            )

    plt.tight_layout(pad=1.1)
    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(); buf.seek(0)
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


def detect_chart_type(prompt: str) -> str:
    p = prompt.lower()
    if "pie"   in p: return "pie"
    if "line"  in p: return "line"
    if any(w in p for w in ["metric", "kpi"]): return "metrics"
    return "bar"


def extract_chart_data(agent_result: dict, prompt: str, project_key: str) -> Optional[Dict]:
    if "chart_data" in agent_result:
        cd = agent_result["chart_data"]
        return {"title": cd.get("title", "Chart"), "labels": cd.get("labels", []), "values": cd.get("values", [])}

    for key, title_template in [
        ("status_counts",   f"{project_key} — Issue Status"),
        ("assignee_counts", f"{project_key} — Issues by Assignee"),
        ("priority_counts", f"{project_key} — Issues by Priority"),
        ("type_counts",     f"{project_key} — Issues by Type"),
    ]:
        if key in agent_result and agent_result[key]:
            p = prompt.lower()
            selected_key = key
            if "assignee" in p or "who" in p:     selected_key = "assignee_counts"
            elif "priority" in p:                  selected_key = "priority_counts"
            elif "type" in p or "kind" in p:       selected_key = "type_counts"
            else:                                  selected_key = "status_counts"
            counts = agent_result.get(selected_key, agent_result[key])
            return {"title": title_template, "labels": list(counts.keys()), "values": list(counts.values())}
    return None


def ensure_jira_chart_data_for_export(prompt: str, project_key: str, agent_result: dict) -> Optional[Dict]:
    """
    For exports, ensure we fetch REAL Jira data.
    Tries multiple strategies to get real data instead of dummy data.
    """
    if not project_key:
        return None
    
    # Try to extract from agent result first
    chart_data = extract_chart_data(agent_result, prompt, project_key)
    if chart_data:
        logger.info(f"[EXPORT] Using Jira data from agent result")
        return chart_data
    
    # If not found, call project summary directly
    try:
        auth_context = get_mcp_auth_context()
        if auth_context:
            summary = call_jira_tool(
                "jira_get_project_summary",
                {
                    "auth_context": auth_context,
                    "project_key": project_key,
                },
            )
            logger.info(f"[EXPORT] Fetched project summary for {project_key}")
            chart_data = extract_chart_data(summary, prompt, project_key)
            if chart_data:
                logger.info(f"[EXPORT] Using Jira data from project summary")
                return chart_data
    except Exception as e:
        logger.warning(f"[EXPORT] Failed to fetch project summary: {e}")
    
    # Alternative: search for issues and aggregate by status
    try:
        if get_jira_auth():
            data = jira_get_cached("search/jql", params={
                "jql": f"project={project_key} ORDER BY created DESC",
                "maxResults": 100,
                "fields": "status,priority,assignee,issuetype",
            })
            logger.info(f"[EXPORT] Fetched {data.get('total', 0)} issues from Jira")
            
            # Aggregate by status
            status_counts = {}
            for issue in data.get("issues", []):
                status_name = issue.get("fields", {}).get("status", {}).get("name", "Unknown")
                status_counts[status_name] = status_counts.get(status_name, 0) + 1
            
            if status_counts:
                logger.info(f"[EXPORT] Using aggregated status data: {status_counts}")
                return {
                    "title": f"{project_key} — Issue Status",
                    "labels": list(status_counts.keys()),
                    "values": list(status_counts.values())
                }
    except Exception as e:
        logger.warning(f"[EXPORT] Failed to aggregate issues by status: {e}")
    
    logger.warning(f"[EXPORT] No real Jira data found for {project_key}")
    return None


# ══════════════════════════════════════════════════════════════════════════════
# Excel Export with Chart
# ══════════════════════════════════════════════════════════════════════════════

def create_excel_with_chart(chart_data: dict, chart_config: dict, project_key: str = "") -> BytesIO:
    """
    Create an Excel file with:
    - Sheet 1: Aggregated data from chart (status counts, etc.)
    - Sheet 2: Detailed Jira issues
    - Sheet 3: Chart image embedded
    """
    from openpyxl.styles import Font, Alignment
    
    # Create workbook
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Add summary headers
    ws_summary['A1'] = "Category"
    ws_summary['B1'] = "Count"
    ws_summary['A1'].font = Font(bold=True, size=12)
    ws_summary['B1'].font = Font(bold=True, size=12)
    ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary['B1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add summary data rows
    labels = chart_data.get("labels", [])
    values = chart_data.get("values", [])
    for idx, (label, value) in enumerate(zip(labels, values), start=2):
        ws_summary[f'A{idx}'] = label
        ws_summary[f'B{idx}'] = value
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    # Add detailed Jira issues sheet
    ws_issues = wb.create_sheet("Jira Issues")
    ws_issues['A1'] = "Issue Key"
    ws_issues['B1'] = "Summary"
    ws_issues['C1'] = "Status"
    ws_issues['D1'] = "Assignee"
    ws_issues['E1'] = "Priority"
    ws_issues['F1'] = "Type"
    ws_issues['G1'] = "Created"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_issues[f'{col}1'].font = Font(bold=True, size=11)
        ws_issues[f'{col}1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Fetch detailed issues from Jira
    try:
        if project_key and get_jira_auth():
            data = jira_get_cached("search/jql", params={
                "jql": f"project={project_key} ORDER BY created DESC",
                "maxResults": 50,
                "fields": "key,summary,status,assignee,priority,issuetype,created",
            })
            
            row = 2
            for issue in data.get("issues", []):
                fields = issue.get("fields", {})
                ws_issues[f'A{row}'] = issue.get("key", "")
                ws_issues[f'B{row}'] = fields.get("summary", "")
                ws_issues[f'C{row}'] = fields.get("status", {}).get("name", "")
                ws_issues[f'D{row}'] = fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned"
                ws_issues[f'E{row}'] = fields.get("priority", {}).get("name", "") if fields.get("priority") else ""
                ws_issues[f'F{row}'] = fields.get("issuetype", {}).get("name", "") if fields.get("issuetype") else ""
                ws_issues[f'G{row}'] = fields.get("created", "")[:10] if fields.get("created") else ""
                row += 1
            
            logger.info(f"[EXCEL] Added {row - 2} detailed Jira issues for {project_key}")
    except Exception as e:
        logger.warning(f"Could not fetch detailed Jira issues for {project_key}: {e}")
    
    # Adjust column widths for issues sheet
    ws_issues.column_dimensions['A'].width = 12
    ws_issues.column_dimensions['B'].width = 35
    ws_issues.column_dimensions['C'].width = 15
    ws_issues.column_dimensions['D'].width = 20
    ws_issues.column_dimensions['E'].width = 12
    ws_issues.column_dimensions['F'].width = 12
    ws_issues.column_dimensions['G'].width = 12
    
    # Add chart image to third sheet
    if chart_config.get("image_base64"):
        ws_chart = wb.create_sheet("Chart")
        
        try:
            # Decode base64 image
            img_b64 = chart_config["image_base64"]
            if img_b64.startswith("data:image/png;base64,"):
                img_b64 = img_b64.replace("data:image/png;base64,", "")
            
            img_data = BytesIO(base64.b64decode(img_b64))
            img_data.seek(0)
            
            # Add image to worksheet
            img = XLImage(img_data)
            img.width = 600
            img.height = 400
            ws_chart.add_image(img, "A1")
            logger.info(f"Chart image embedded in Excel export")
        except Exception as e:
            logger.warning(f"Could not embed chart image in Excel: {e}")
    
    # Save to BytesIO
    excel_buf = BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)
    return excel_buf


# ══════════════════════════════════════════════════════════════════════════════
# Main process function
# ══════════════════════════════════════════════════════════════════════════════

def process_prompt(prompt: str) -> Dict:
    prompt_lower = prompt.lower().strip()
    project_key  = get_user_project_key()

    if not project_key:
        return {"error": "No project selected. Please select a Jira project first.", "needs_project": True}

    # ── Special: clear all dashboards ────────────────────────────────────────
    if any(w in prompt_lower for w in ["clear all", "clear dashboard", "remove all", "delete all"]):
        uid = user_id()
        if uid and uid in _user_dashboards:
            _user_dashboards[uid] = {}
        return {"response": "All dashboards cleared! ✓", "cleared": True}

    chart_keywords  = ["chart", "graph", "pie", "bar", "line", "visuali", "plot", "metrics", "kpi"]
    create_keywords = ["create task", "create a task", "new task", "add task",
                       "create ticket", "open a ticket", "raise a ticket"]
    export_keywords = ["export", "excel", "download", "xlsx", "csv"]

    wants_chart  = any(w in prompt_lower for w in chart_keywords)
    wants_create = any(p in prompt_lower for p in create_keywords)
    wants_export = any(w in prompt_lower for w in export_keywords)
    chart_type   = detect_chart_type(prompt)
    jira_ok      = get_jira_auth() is not None

    # Fast dispatch for non-chart queries
    if jira_ok and not wants_chart and not wants_create:
        fast = fast_dispatch(prompt_lower, project_key)
        if fast:
            logger.info("[FAST DISPATCH] LLM skipped")
            return fast

    # Run MCP agent
    agent_result = {}
    if jira_ok:
        agent_result = run_jira_agent(prompt, project_key)

    # Chart path
    if wants_chart:
        # For exports, prioritize real Jira data
        if wants_export:
            chart_data = ensure_jira_chart_data_for_export(prompt, project_key, agent_result)
        else:
            chart_data = extract_chart_data(agent_result, prompt, project_key)

        if not chart_data and jira_ok:
            try:
                summary = call_jira_tool(
                    "jira_get_project_summary",
                    {
                        "auth_context": get_mcp_auth_context() or {},
                        "project_key": project_key,
                    },
                )
                chart_data = extract_chart_data(summary, prompt, project_key)
            except Exception as e:
                logger.warning(f"Project summary fallback failed: {e}")

        if not chart_data and llm:
            try:
                system = SystemMessage(content='Generate chart data as JSON only:\n{"title":"...","labels":[...],"values":[...]}')
                resp   = llm.invoke([system, HumanMessage(content=prompt)])
                match  = re.search(r"\{.*\}", resp.content, re.DOTALL)
                if match:
                    chart_data = json.loads(match.group())
            except Exception:
                pass

        if not chart_data:
            logger.warning(f"[CHART] Using dummy data - no real Jira data available")
            chart_data = {"title": "Sample Distribution",
                          "labels": ["A", "B", "C", "D"], "values": [35, 25, 25, 15]}

        chart_id = str(uuid.uuid4())[:8]

        # Auto-switch: bar/line with only 1 data point looks wrong → use pie or metrics
        effective_type = chart_type
        if len(chart_data["labels"]) == 1 and chart_type in ("bar", "line"):
            effective_type = "metrics"
            logger.info(f"Auto-switched chart type to metrics (only 1 data point)")

        image_b64 = generate_chart_b64(effective_type, chart_data["title"],
                                        chart_data["labels"], chart_data["values"])
        chart_type = effective_type
        chart_config = {
            "chart_id":     chart_id,
            "chart_type":   chart_type,
            "title":        chart_data["title"],
            "labels":       chart_data["labels"],
            "values":       chart_data["values"],
            "image_base64": image_b64,
            "created_at":   datetime.now().isoformat(),
            "data_source":  "jira_mcp" if jira_ok else "generated",
        }

        result = {"response": f"{chart_type.upper()} chart generated.", "chart_config": chart_config}

        if wants_export:
            logger.info(f"[EXPORT] Creating Excel with chart data: {len(chart_data.get('labels', []))} items")
            excel_buf = create_excel_with_chart(chart_data, chart_config, project_key)
            chart_config["export_format"] = "excel"
            chart_config["export_data"]   = base64.b64encode(excel_buf.getvalue()).decode()
            result["auto_export"]  = True
            result["dashboard_id"] = chart_id
            result["data_source"]  = chart_config.get("data_source", "generated")
            logger.info(f"[EXPORT] Excel file generated successfully (dashboard_id: {chart_id})")

        return result

    # Create task path
    if wants_create:
        jira_task = agent_result.get("key") or agent_result.get("task_id")
        if jira_task:
            cfg    = get_jira_auth()
            server = cfg["base_url"] if cfg else ""
            if cfg and cfg["auth_type"] == "oauth":
                site   = session.get("jira_site_name", "")
                server = f"https://{site}.atlassian.net" if site else server
            return {
                "response":  f"Jira task created: {jira_task}",
                "jira_task": {"task_id": jira_task, "url": f"{server}/browse/{jira_task}", "summary": prompt[:100]},
            }

    # Query / text path
    if agent_result.get("result_type"):
        return agent_result

    issues = agent_result.get("issues", [])
    if issues:
        return {"response": f"Found {len(issues)} issue(s)", "result_type": "issues", "jira_results": issues}

    raw = agent_result.get("raw_response", "")
    if raw:
        return {"response": raw}

    return {"response": "Request processed."}


# ══════════════════════════════════════════════════════════════════════════════
# Login / OAuth Routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/login")
def login_page():
    """Show login page to unauthenticated users."""
    if is_authenticated():
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/auth/login")
def auth_login():
    """Start Atlassian OAuth flow — manual state to avoid CSRF issues."""
    import secrets
    state        = secrets.token_urlsafe(32)
    redirect_uri = url_for("auth_callback", _external=True)

    # Store state in a separate simple cookie (more reliable than session for OAuth)
    # prompt parameter — only passed when explicitly requested (e.g. after logout)
    prompt_param = f"&prompt={request.args.get('prompt', '')}" if request.args.get('prompt') else ""

    response = redirect(
        f"https://auth.atlassian.com/authorize"
        f"?audience=api.atlassian.com"
        f"&client_id={os.getenv('JIRA_CLIENT_ID')}"
        f"&scope={requests.utils.quote(JIRA_OAUTH_SCOPES, safe='')}"
        f"&redirect_uri={requests.utils.quote(redirect_uri, safe='')}"
        f"&state={state}"
        f"&response_type=code"
        f"{prompt_param}"
    )
    response.set_cookie("oauth_state", state, max_age=600, httponly=True, samesite="Lax")
    return response


@app.route("/auth/callback")
def auth_callback():
    """Handle OAuth callback — exchange code for token manually."""
    try:
        # Verify state from cookie (not session — more reliable)
        stored_state   = request.cookies.get("oauth_state", "")
        returned_state = request.args.get("state", "")

        if not stored_state or stored_state != returned_state:
            logger.error(f"State mismatch: stored={stored_state[:10]}... returned={returned_state[:10]}...")
            return render_template("login.html",
                error="Login session expired. Please try again.")

        # Exchange code for token manually
        code         = request.args.get("code")
        redirect_uri = url_for("auth_callback", _external=True)

        token_resp = requests.post(
            "https://auth.atlassian.com/oauth/token",
            json={
                "grant_type":    "authorization_code",
                "client_id":     os.getenv("JIRA_CLIENT_ID"),
                "client_secret": os.getenv("JIRA_CLIENT_SECRET"),
                "code":          code,
                "redirect_uri":  redirect_uri,
            },
            headers={"Content-Type": "application/json"}
        )
        token_resp.raise_for_status()
        token        = token_resp.json()
        access_token = token.get("access_token")

        # Fetch all accessible Jira sites first (always works)
        sites = fetch_accessible_sites(access_token)

        if not sites:
            return render_template("login.html",
                error="No Jira sites found for your account. Make sure you have access to at least one Jira project.")

        # Fetch user info — try /me first, fall back to Jira myself endpoint
        u = {}
        try:
            user_resp = requests.get(
                "https://api.atlassian.com/me",
                headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
            )
            user_resp.raise_for_status()
            u = user_resp.json()
            logger.info(f"User info from /me: {u.get('email', 'unknown')}")
        except Exception as me_err:
            logger.warning(f"/me endpoint failed ({me_err}), trying Jira myself endpoint")
            try:
                # Try getting user info from Jira API itself
                cloud_id = sites[0]["id"]
                myself_resp = requests.get(
                    f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/api/3/myself",
                    headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
                )
                myself_resp.raise_for_status()
                jira_user = myself_resp.json()
                u = {
                    "name":    jira_user.get("displayName", ""),
                    "email":   jira_user.get("emailAddress", ""),
                    "picture": jira_user.get("avatarUrls", {}).get("48x48", ""),
                }
                logger.info(f"User info from Jira myself: {u.get('email', 'unknown')}")
            except Exception as jira_err:
                logger.warning(f"Jira myself also failed ({jira_err}), using site name as identity")
                u = {
                    "name":    sites[0].get("name", "Jira User"),
                    "email":   "",
                    "picture": "",
                }

        session.permanent = True
        session["oauth_token"] = token
        session["user"] = {
            "name":   u.get("name", u.get("displayName", "")),
            "email":  u.get("email", ""),
            "avatar": u.get("picture", ""),
        }
        session["jira_sites"] = sites

        # If only one site, auto-select it
        if len(sites) == 1:
            session["jira_cloud_id"]  = sites[0]["id"]
            session["jira_site_name"] = sites[0]["name"]
            session["jira_site_url"]  = sites[0].get("url", "")
            ensure_selected_project()
            return redirect(url_for("index"))

        # Multiple sites — let user pick
        return redirect(url_for("select_site"))

    except Exception as e:
        logger.error(f"OAuth callback error: {e}")
        return render_template("login.html", error=f"Login failed: {str(e)}")


@app.route("/select-site")
@login_required
def select_site():
    """Let user pick which Jira site to use (if they have multiple)."""
    sites = session.get("jira_sites", [])
    return render_template("select_site.html", sites=sites, user=current_user())


@app.route("/select-site", methods=["POST"])
@login_required
def select_site_post():
    """Save selected Jira site."""
    cloud_id  = request.form.get("cloud_id")
    sites     = session.get("jira_sites", [])
    site      = next((s for s in sites if s["id"] == cloud_id), None)
    if not site:
        return redirect(url_for("select_site"))

    session["jira_cloud_id"]  = cloud_id
    session["jira_site_name"] = site["name"]
    session["jira_site_url"]  = site.get("url", "")
    ensure_selected_project()
    return redirect(url_for("index"))


@app.route("/select-project")
@login_required
def select_project():
    """Let user pick which Jira project to work with."""
    token    = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    projects = fetch_user_projects(cloud_id, token.get("access_token", ""))
    return render_template("select_project.html", projects=projects,
                           user=current_user(), site=session.get("jira_site_name"))


@app.route("/select-project", methods=["POST"])
@login_required
def select_project_post():
    """Save selected project."""
    project_key  = request.form.get("project_key", "").upper().strip()
    project_name = request.form.get("project_name", "")
    if not project_key:
        return redirect(url_for("select_project"))
    session["project_key"]  = project_key
    session["project_name"] = project_name
    return redirect(url_for("index"))


@app.route("/api/project/select", methods=["POST"])
@login_required
def select_project_api():
    payload = request.get_json() or {}
    requested_key = (payload.get("project_key") or "").upper().strip()
    if not requested_key:
        return jsonify({"error": "project_key required"}), 400

    token = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    projects = fetch_user_projects(cloud_id, token.get("access_token", ""))
    selected = next((p for p in projects if (p.get("key") or "").upper() == requested_key), None)
    if not selected:
        return jsonify({"error": f"Project {requested_key} not found or no access"}), 404

    session["project_key"] = selected.get("key", requested_key)
    session["project_name"] = selected.get("name", "")
    clear_user_cache()
    return jsonify({
        "message": "Project selected",
        "project_key": session.get("project_key", ""),
        "project_name": session.get("project_name", ""),
    })


@app.route("/auth/status")
def auth_status():
    if is_authenticated():
        cfg = get_jira_auth()
        return jsonify({
            "connected":    True,
            "auth_type":    cfg["auth_type"] if cfg else "unknown",
            "user":         current_user(),
            "site_name":    session.get("jira_site_name", ""),
            "project_key":  get_user_project_key(),
            "project_name": session.get("project_name", ""),
        })
    return jsonify({"connected": False})


@app.route("/auth/logout")
def auth_logout():
    # Clear per-user dashboards from memory on logout
    uid = user_id()
    if uid and uid in _user_dashboards:
        del _user_dashboards[uid]
        logger.info(f"Dashboards cleared for user {uid}")
    # Clear user cache
    clear_user_cache()
    session.clear()
    response = redirect(url_for("login_page"))
    response.delete_cookie("oauth_state")
    return response


# ══════════════════════════════════════════════════════════════════════════════
# App Routes (all protected)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def index():
    if session.get("jira_cloud_id") and not get_user_project_key():
        ensure_selected_project()
    return render_template("index.html",
        user=current_user(),
        site=session.get("jira_site_name", ""),
        project_key=get_user_project_key(),
        project_name=session.get("project_name", ""),
    )


@app.route("/api/generate", methods=["POST"])
@login_required
def generate_dashboard():
    data   = request.get_json()
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return jsonify({"error": "Prompt is required"}), 400

    result = process_prompt(prompt)

    if result.get("error") and not result.get("needs_project"):
        return jsonify(result), 400

    if result.get("chart_config"):
        dash_id = result["chart_config"]["chart_id"]
        save_user_dashboard(dash_id, {
            "id":           dash_id,
            "prompt":       prompt,
            "chart_config": result["chart_config"],
            "jira_task":    result.get("jira_task"),
            "created_at":   datetime.now().isoformat(),
        })
        result["dashboard_id"] = dash_id

    return jsonify(result)


@app.route("/api/dashboards", methods=["GET"])
@login_required
def list_dashboards():
    dashes = get_user_dashboards()
    return jsonify({"dashboards": list(dashes.values()), "count": len(dashes)})


@app.route("/api/dashboard/<dashboard_id>", methods=["GET"])
@login_required
def get_dashboard(dashboard_id):
    dashes = get_user_dashboards()
    if dashboard_id not in dashes:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dashes[dashboard_id])


@app.route("/api/dashboard/<dashboard_id>", methods=["PUT"])
@login_required
def update_dashboard(dashboard_id):
    dashes = get_user_dashboards()
    if dashboard_id not in dashes:
        return jsonify({"error": "Not found"}), 404
    prompt = (request.get_json() or {}).get("prompt", "").strip()
    if not prompt:
        return jsonify({"error": "Prompt required"}), 400
    result = process_prompt(prompt)
    dashes[dashboard_id].update({"prompt": prompt, "chart_config": result.get("chart_config"),
                                  "updated_at": datetime.now().isoformat()})
    save_user_dashboard(dashboard_id, dashes[dashboard_id])
    return jsonify({"message": "Updated", "dashboard": dashes[dashboard_id]})


@app.route("/api/jira/tasks", methods=["GET"])
@login_required
def get_jira_tasks():
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503
    project_key = get_user_project_key()
    try:
        data   = jira_get_cached("search/jql", params={
            "jql":        f"project={project_key} ORDER BY created DESC",
            "maxResults": 50,
            "fields":     "summary,status,assignee,priority,created",
        })
        cfg    = get_jira_auth()
        site   = session.get("jira_site_name", "")
        server = f"https://{site}.atlassian.net" if site else os.getenv("JIRA_SERVER", "")
        tasks  = []
        for issue in data.get("issues", []):
            f = issue["fields"]
            tasks.append({
                "task_id":  issue["key"],
                "summary":  f.get("summary", ""),
                "status":   f["status"]["name"],
                "assignee": f["assignee"]["displayName"] if f.get("assignee") else "Unassigned",
                "priority": f["priority"]["name"] if f.get("priority") else "None",
                "created":  f.get("created", ""),
                "url":      f"{server}/browse/{issue['key']}",
            })
        return jsonify({"tasks": tasks, "count": len(tasks)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/projects", methods=["GET"])
@login_required
def get_user_projects_api():
    """Return list of projects for the current user's Jira site."""
    token    = session.get("oauth_token", {})
    cloud_id = session.get("jira_cloud_id")
    if not cloud_id:
        return jsonify({"error": "No Jira site selected"}), 400
    projects = fetch_user_projects(cloud_id, token.get("access_token", ""))
    return jsonify({"projects": projects})


@app.route("/api/jira/tasks", methods=["POST"])
@login_required
def create_jira_task():
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503
    project_key = get_user_project_key()
    data        = request.get_json()
    try:
        payload = {"fields": {
            "project":     {"key": project_key},
            "summary":     data.get("summary", "New Task"),
            "issuetype":   {"name": "Task"},
            "priority":    {"name": data.get("priority", "Medium")},
            "description": {"type": "doc", "version": 1, "content": [
                {"type": "paragraph", "content": [{"type": "text", "text": data.get("description", "")}]}
            ]},
        }}
        result    = jira_post("issue", payload)
        clear_user_cache()
        issue_key = result.get("key")
        site      = session.get("jira_site_name", "")
        server    = f"https://{site}.atlassian.net" if site else os.getenv("JIRA_SERVER", "")
        return jsonify({"message": "Task created", "task": {"task_id": issue_key, "url": f"{server}/browse/{issue_key}"}}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/tasks/<task_id>/edit-data", methods=["GET"])
@login_required
def get_jira_task_edit_data(task_id):
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503

    try:
        issue = jira_get(f"issue/{task_id}", params={"fields": "summary,status,assignee,priority"})
        fields = issue.get("fields", {})

        current = {
            "summary": fields.get("summary", ""),
            "status": fields.get("status", {}).get("name", "") if fields.get("status") else "",
            "assignee": {
                "accountId": fields.get("assignee", {}).get("accountId", "") if fields.get("assignee") else "",
                "displayName": fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned",
            },
            "priority": fields.get("priority", {}).get("name", "") if fields.get("priority") else "",
        }

        transitions_data = jira_get(f"issue/{task_id}/transitions")
        transitions = [
            {"id": t.get("id", ""), "name": t.get("name", "")}
            for t in transitions_data.get("transitions", [])
        ]

        assignable_raw = jira_get("user/assignable/search", params={"issueKey": task_id, "maxResults": 100})
        assignable_users = assignable_raw if isinstance(assignable_raw, list) else []
        assignees = [
            {
                "accountId": u.get("accountId", ""),
                "displayName": u.get("displayName", "Unknown User"),
            }
            for u in assignable_users
            if u.get("accountId")
        ]

        priority_names = []
        try:
            editmeta = jira_get(f"issue/{task_id}/editmeta")
            allowed_values = (
                editmeta.get("fields", {})
                .get("priority", {})
                .get("allowedValues", [])
            )
            priority_names = [p.get("name", "") for p in allowed_values if p.get("name")]
        except Exception:
            # Keep endpoint resilient even if editmeta is restricted.
            priority_names = []

        if not priority_names:
            priority_names = ["Lowest", "Low", "Medium", "High", "Highest"]

        if current["priority"] and current["priority"] not in priority_names:
            priority_names.insert(0, current["priority"])

        return jsonify({
            "task_id": task_id,
            "current": current,
            "assignees": assignees,
            "priorities": priority_names,
            "statuses": transitions,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jira/tasks/<task_id>", methods=["PUT"])
@login_required
def update_jira_task(task_id):
    if not get_jira_auth():
        return jsonify({"error": "Jira not connected"}), 503
    data = request.get_json()
    try:
        fields = {}
        if "summary" in data:
            fields["summary"] = data.get("summary", "")

        if "assignee_account_id" in data:
            account_id = (data.get("assignee_account_id") or "").strip()
            fields["assignee"] = {"accountId": account_id} if account_id else None

        if "priority" in data:    fields["priority"]    = {"name": data["priority"]}
        if "description" in data: fields["description"] = {"type": "doc", "version": 1, "content": [
            {"type": "paragraph", "content": [{"type": "text", "text": data["description"]}]}]}

        if fields:
            jira_put(f"issue/{task_id}", {"fields": fields})

        if "status_id" in data and data.get("status_id"):
            jira_post(f"issue/{task_id}/transitions", {"transition": {"id": str(data.get("status_id"))}})
        if "status" in data:
            trans   = jira_get(f"issue/{task_id}/transitions")
            matched = next((t for t in trans.get("transitions", []) if t["name"].lower() == data["status"].lower()), None)
            if matched:
                jira_post(f"issue/{task_id}/transitions", {"transition": {"id": matched["id"]}})
        clear_user_cache()
        return jsonify({"message": f"{task_id} updated"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/export/<dashboard_id>", methods=["GET"])
@login_required
def export_dashboard(dashboard_id):
    dashes = get_user_dashboards()
    if dashboard_id not in dashes:
        return jsonify({"error": "Not found"}), 404
    
    dash = dashes[dashboard_id]
    cfg = dash.get("chart_config", {})
    project_key = get_user_project_key()
    
    # Prepare chart data
    chart_data = {
        "labels": cfg.get("labels", []),
        "values": cfg.get("values", []),
    }
    
    # Create Excel with data table and chart image
    excel_buf = create_excel_with_chart(chart_data, cfg, project_key)
    
    return send_file(excel_buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"dashboard_{dashboard_id}.xlsx")


@app.route("/api/metrics", methods=["GET"])
@login_required
def get_metrics():
    dashes = get_user_dashboards()
    types  = ["pie", "bar", "line", "metrics"]
    return jsonify({
        "total_dashboards": len(dashes),
        "chart_types": {t: sum(1 for d in dashes.values() if d.get("chart_config", {}).get("chart_type") == t) for t in types},
    })


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("DEBUG", "true").lower() in {"1", "true", "yes", "on"}
    print("=" * 60)
    print("  Dashboard Generator — Multi-Tenant Edition")
    print("=" * 60)
    print(f"  OAuth Client : {os.getenv('JIRA_CLIENT_ID','Not set')}")
    print(f"  LLM Model    : {os.getenv('QWEN_MODEL','Not set')}")
    print(f"  Login URL    : http://localhost:{port}/login")
    print("=" * 60)
    app.run(debug=debug, host="0.0.0.0", port=port)
